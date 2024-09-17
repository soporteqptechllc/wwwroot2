
# ******************************************************************
# -*- coding: utf-8 -*-
# Consulta Basica de una Base de Datos ya creada en PostgreSQL
# Prog. para 
# Import PostgreSQL
import psycopg2
# Import datetime  
from datetime import date, datetime, timedelta, timezone

# Declaro variables
colu0 = list()
colu1 = list()
colu2 = list()
colu3 = list()
colu4 = list()
listx = ''
listy = ''
fecha = list()
fecha = ['01','01','2023']
year = date.today().year # current year
month = date.today().month # current month
day = date.today().day # current day
week = date.today().isocalendar()[1] # current weeek from year
fecha[0] = day
fecha[1] = month
fecha[2] = year
ano = 2023
# Se crea la funcion de Excel
def excel(fecha):
    
    from openpyxl import load_workbook

    FILE_PATH ='C:/QPTECH/EXCEL/PLANTILLAS/ConsumoDeAguaMolino3_Plantilla.xlsx'
    sheet3 = 'Mensual'
    
    workbook= load_workbook(FILE_PATH,read_only=False)
    ws3=workbook[sheet3]

    # Colocar el año en la Fila 1 Columna A
    ws3.cell(row = 2,column = 1, value = ano)
    # Carga de la Columna B
    rowIndex = 2
    for filas in colu2:
        # Cargo el dato en la Celda
        ws3.cell(row = rowIndex,column = 2, value = filas)
        rowIndex +=1
    # Carga de la Columna C
    rowIndex = 2
    for filas in colu4:
        # Cargo el dato en la Celda
        ws3.cell(row = rowIndex,column = 3, value = filas)
        rowIndex +=1
    # Carga de la Columna D
    rowIndex = 2
    for filas in colu1:
        # Cargo el dato en la Celda
        ws3.cell(row = rowIndex,column = 4, value = filas)
        rowIndex +=1

    fileName = f"C:/QPTECH/EXCEL/AGUA_MOLINO3/ConsumoDeAguaMolino3_{fecha[0]}-{fecha[1]}-{fecha[2]}.xlsx"
    
    workbook.save(fileName)  
    return fileName

# Determinar si el ano es valido
if ano:
    if ano == 1 and year > 2023:
        datoyear = year - 1
    else:
        datoyear = year
else:
    datoyear = year
#print (datoyear)
# Conexion a tu base de datos
# Usa el método conectar()
try:    
    credenciales = {
    "dbname": "cemento",
    "user": "mantenimiento",
    "password": "#Argos2017",
    "host": "localhost",
    "port": ''
    }
    cnxn1 = psycopg2.connect(**credenciales)   
    cursor1 = cnxn1.cursor()

    #print("Iniciado PostgreSQL.....")

    # Determinar el ano actual
    now = datetime.now() # current date and time
    year = int(now.strftime("%Y")) # current year)
    print(year)

    # Crear el texto del Query   
    texto_query1 = """
    SELECT year, MAX(month), MAX(fecha), MAX(volumen) FROM public."ViewAguaLanzasM3Month" 
    WHERE year = %s
    GROUP BY year
    """ % (datoyear-1) 
    # Ejecutar una consulta
    cursor1.execute(texto_query1)
    # Se crea la costante del acumulado del ano anterior
    acum = 0  
    while 1:
        row = cursor1.fetchone()
        #print(row)
        if not row:
            break      
        #print(int(row[3]))
        acum = int(row[3])
    #print(acum)
    # Crear el texto del Query   
    texto_query2 = """
    SELECT * FROM public."ViewAguaLanzasM3Month" 
    WHERE year = %s
    """ % (datoyear)
    # Ejecutar una consulta
    cursor1.execute(texto_query2)
    
    # Se crea la lista de los datos
    ctto = 0
    dato = 0  
    while 1:
        row = cursor1.fetchone()
        ctto = ctto + 1
        if not row:
            break
        if ctto == 1:
            dato = acum
            colu0.append(int(row[0])) # year
            if row[1] == 1:
                colu1.append('ENERO') # month
            elif row[1] == 2:
                colu1.append('FEBRERO')
            elif row[1] == 3:
                colu1.append('MARZO')
            elif row[1] == 4:
                colu1.append('ABRIL')
            elif row[1] == 5:
                colu1.append('MAYO')
            elif row[1] == 6:
                colu1.append('JUNIO')
            elif row[1] == 7:
                colu1.append('JULIO')
            elif row[1] == 8:
                colu1.append('AGOSTO')
            elif row[1] == 9:
                colu1.append('SEPTIEMBRE')
            elif row[1] == 10:
                colu1.append('OCTUBRE')
            elif row[1] == 11:
                colu1.append('NOVIEMBRE')
            elif row[1] == 12:
                colu1.append('DICIEMBRE')
            colu2.append(row[2].strftime("%d-%m-%Y")) # datetime
            colu3.append(int(row[3])) # volumen
            colu4.append(int(row[3]) - dato) # acumulado
            dato = int(row[3]) 
        else:
            colu0.append(int(row[0])) # year
            if row[1] == 1:
                colu1.append('ENERO') # month
            elif row[1] == 2:
                colu1.append('FEBRERO')
            elif row[1] == 3:
                colu1.append('MARZO')
            elif row[1] == 4:
                colu1.append('ABRIL')
            elif row[1] == 5:
                colu1.append('MAYO')
            elif row[1] == 6:
                colu1.append('JUNIO')
            elif row[1] == 7:
                colu1.append('JULIO')
            elif row[1] == 8:
                colu1.append('AGOSTO')
            elif row[1] == 9:
                colu1.append('SEPTIEMBRE')
            elif row[1] == 10:
                colu1.append('OCTUBRE')
            elif row[1] == 11:
                colu1.append('NOVIEMBRE')
            elif row[1] == 12:
                colu1.append('DICIEMBRE')
            colu2.append(row[2].strftime("%d-%m-%Y")) # datetime
            colu3.append(int(row[3])) # volumen
            colu4.append(int(row[3]) - dato) # acumulado
            dato = int(row[3])
        #print(ctto)
        #print(dato)
        #print(int(row[0]))
        #print(int(row[1]))   
        #print(int(row[3]))
        #print(row[2])
    print(colu0,colu1,colu2,colu3,colu4)
    # Creo las listas X y Y
    for i in reversed(colu1):
        if listx == '':
            listx = str(i)
        else:
            listx = str(i) + ',' + listx

    for i in reversed(colu4):
        if listy == '':
            listy = str(i)
        else:
            listy = str(i) + ',' + listy
    #print(listx)
    #print(listy)

    # Ejecuto el Metodo de carga en Excel    
    print("Iniciado Escritura de EXCEL.....")
    excel(fecha)
except Exception as e: 
    #print("Error de conexión PostgreSQL: ", e)
    cursor1.close()  
    cnxn1.close()
finally:
# Se cierra la conexion SQL y la conexion al Servidos
    cursor1.close()  
    cnxn1.close()
    #print("La conexión PostgreSQL está cerrada")


