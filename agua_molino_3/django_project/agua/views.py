# Import PostgreSQL
import psycopg2
# Import pytz
import pytz
# Import call
from subprocess import call
# Import send_mail  
from send_mail import mail
# Import REST FRAMEWORK
from rest_framework.generics import (
    ListAPIView,
    CreateAPIView,
    RetrieveAPIView,
    DestroyAPIView,
    UpdateAPIView,
    RetrieveUpdateAPIView
)
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework import status
from rest_framework.response import Response
# Import Django
from django.shortcuts import render
from django.contrib import messages
from django.views.generic import (
    View, 
    ListView, 
    DetailView,
    CreateView,
    TemplateView,
    UpdateView,
    DeleteView
    )
from django.contrib.auth.models import User
# Local Serializers
from .serializers import (
    MedidorAguaRioSerializer,
    MedidorAguaRioSerializer1,
    MedidorAguaLanzasMolino3Serializer,
    MedidorAguaLanzasMolino3Serializer1
    )
# Import datetime  
from datetime import date, datetime, timedelta, timezone

# Local Models de la BD
from .models import (
    MedidorAguaRio
    )
# Import send_mail  
from send_mail import mail
# Import tkinter 
from tkinter import ttk, Tk

# Create your views here.

# ********************************************************************************************
# ********************************************************************************************
# Lista todas las mediciones de agua de rio de la empresa en una HTML
class ListMedicionesAguaRio(ListView):
    template_name = 'dashboard/mediationlist.html'
    paginate_by = 100
    ordering = '-created_at'
    
    def get_queryset(self):
    # Codigo que filtra lista de mediciones desde la pantalla
    # El 'self.request.GET.get' nos permite recoger una variable desde el HTML
    # desde un <input> y un <button>
        clave = self.request.GET.get('kword', '')
        lista = MedidorAguaRio.objects.filter(
            created_at__icontains=clave).order_by('-created_at')
        return lista

# List View from Mediciones de Agua de Rio in an Url
class ListMedicionesAguaRioAPI(ListAPIView):
    # Paquete que desencripta y autentifica el usuario atreves de un token
    authentication_classes = (TokenAuthentication,)
    permission_classes = [IsAuthenticated,]
    serializer_class = MedidorAguaRioSerializer

    def get_queryset(self):
        # Recuperar usuario
        usuario = self.request.user
        # Valido si esta activo el usuario
        if usuario.is_active:
            return MedidorAguaRio.objects.all()
        else:
            return Response({'Error':'Este usuario no puede iniciar sesion'}, status = status.HTTP_401_UNAUTHORIZED)  

# Create from Mediciones de Agua de Rio in an Url
class CrearMedicionesAguaRioAPI(CreateAPIView):
    # Paquete que desencripta y autentifica el usuario atreves de un token
    authentication_classes = (TokenAuthentication,)
    permission_classes = [IsAuthenticated,]    
    serializer_class = MedidorAguaRioSerializer1
    def get_queryset(self):
        # Recuperar usuario
        usuario = self.request.user
        # Valido si esta activo el usuario
        if usuario.is_active:
            return Response({'mensaje':'Todo Ok.'}, status = status.HTTP_200_OK)
        else:
            return Response({'Error':'Este usuario no puede iniciar sesion'}, status = status.HTTP_401_UNAUTHORIZED)
# VISTA GraficaMedicionAguaRio1 ========================================================================
class GraficaMedicionAguaRio1(View):
    # Para all() : .exists(),.count(),.first(),.last(),.astimezone(tz)
    def get(self, request):
        # Declaracion de variables
        tz = pytz.timezone('America/Bogota') # current time zone local
        now = datetime.now() # current date and time
        hoy = now.strftime("%Y-%m-%d") # current day
        limite = now - timedelta(days=45)  # day subtraction operation
        DataMedicion = MedidorAguaRio.objects.filter(created_at__gte=limite).order_by('-created_at')
        listx = ''
        listy = ''

        for i in DataMedicion:
            if listx == '':
                listx = str(i.created_at.astimezone(tz).strftime("%Y-%m-%d  %H:%M"))
                listy = str(i.volumen)
            else:              
                listx = str(i.created_at.astimezone(tz).strftime("%Y-%m-%d  %H:%M")) + ',' + listx
                listy = str(i.volumen) + ',' + listy
        
        messages.info(
            request, '  Por favor espere que carge la información')
        greeting = {}
        greeting['heading'] = "Vista Graf.Volumen"
        greeting['pageview'] = "Sistema de agua recuperada"
        greeting['listaX'] = listx
        greeting['listaY'] = listy
        greeting['tituloX'] = "Volumen del agua de las Bombas de rio (m3)"
        greeting['titulotxt1'] = "Nivel del Tanque (%)"
        greeting['titulotxt2'] = "Tanque de Agua Cruda"
        greeting['volumentxt'] = DataMedicion.first().volumen
        greeting['flujotxt'] = DataMedicion.first().flujo
        greeting['niveltxt'] = DataMedicion.first().nivel
        greeting['bomba1txt'] = DataMedicion.first().bomba1
        greeting['bomba2txt'] = DataMedicion.first().bomba2
        greeting['fechabdtxt'] = str(DataMedicion.first().created_at.astimezone(tz).strftime("%Y-%m-%d  %H:%M:%S"))
        greeting['fechaplctxt'] = str(DataMedicion.first().datatime.astimezone(tz).strftime("%Y-%m-%d  %H:%M:%S"))
        return render (request,'dashboard/grafica1.html',greeting)

# VISTA GraficaMedicionAguaRio2 ========================================================================
class GraficaMedicionAguaRio2(View):
    def get(self, request):
        # Declaracion de variables
        tz = pytz.timezone('America/Bogota') # current time zone local
        now = datetime.now() # current date and time
        hoy = now.strftime("%Y-%m-%d") # current day
        limite = now - timedelta(days=45)  # day subtraction operation
        DataMedicion = MedidorAguaRio.objects.filter(created_at__gte=limite).order_by('-created_at')
        listx = ''
        listy = ''

        for i in DataMedicion:
            if listx == '':
                listx = str(i.created_at.astimezone(tz).strftime("%Y-%m-%d  %H:%M"))

                listy = str(i.flujo)
            else:
                listx = str(i.created_at.astimezone(tz).strftime("%Y-%m-%d  %H:%M")) + ',' + listx
                listy = str(i.flujo) + ',' + listy

        messages.info(
            request, '  Por favor espere que carge la información')
        greeting = {}
        greeting['heading'] = "Vista Graf.Flujo"
        greeting['pageview'] = "Sistema de agua recuperada"
        greeting['listaX'] = listx
        greeting['listaY'] = listy
        greeting['tituloX'] = "Flujo del agua de las Bombas de rio (G/m)"
        greeting['titulotxt1'] = "Nivel del Tanque (%)"
        greeting['titulotxt2'] = "Tanque de Agua Cruda"
        greeting['volumentxt'] = DataMedicion.first().volumen
        greeting['flujotxt'] = DataMedicion.first().flujo
        greeting['niveltxt'] = DataMedicion.first().nivel
        greeting['bomba1txt'] = DataMedicion.first().bomba1
        greeting['bomba2txt'] = DataMedicion.first().bomba2
        greeting['fechabdtxt'] = str(
            DataMedicion.first().created_at.astimezone(tz).strftime("%Y-%m-%d  %H:%M:%S"))       
        greeting['fechaplctxt'] = str(
            DataMedicion.first().datatime.astimezone(tz).strftime("%Y-%m-%d  %H:%M:%S"))
        return render(request, 'dashboard/grafica2.html', greeting)
# ********************************************************************************************
# ********************************************************************************************
# Create from Mediciones de Agua de Lanzas de Molino3 in an Url
class CrearMedidorAguaLanzasMolino3API(CreateAPIView):
    # Paquete que desencripta y autentifica el usuario atreves de un token
    authentication_classes = (TokenAuthentication,)
    permission_classes = [IsAuthenticated,]    
    serializer_class = MedidorAguaLanzasMolino3Serializer1
    def get_queryset(self):
        # Recuperar usuario
        usuario = self.request.user
        # Valido si esta activo el usuario
        if usuario.is_active:
            return Response({'mensaje':'Todo Ok.'}, status = status.HTTP_200_OK)
        else:
            return Response({'Error':'Este usuario no puede iniciar sesion'}, status = status.HTTP_401_UNAUTHORIZED)


# ********************************************************************************************
# ********************************************************************************************
# VISTA GraficaConsumoAguaLanzasM3 ANUAL =======================================================================
class GraficaConsumoAguaLanzasM3A(View):
    
    def get(self, request, ano = None, email = None):
        # Declaro variables
        colu0 = list()
        colu1 = list()
        colu2 = list()
        colu3 = list()
        colu4 = list()
        listx = ''
        listy = ''
        fecha = list()
        fecha = ['01','01','2023']
        env_email = 0
        year = date.today().year # current year
        month = date.today().month # current month
        day = date.today().day # current day
        week = date.today().isocalendar()[1] # current weeek from year
        fecha[0] = day
        fecha[1] = month
        fecha[2] = year
        file_name = str(f"ConsumoDeAguaMolino3_{fecha[0]}-{fecha[1]}-{fecha[2]}.xlsx")
        file_path = str(f"C:/QPTECH/EXCEL/AGUA_MOLINO3/ConsumoDeAguaMolino3_{fecha[0]}-{fecha[1]}-{fecha[2]}.xlsx")
        # Determinar si el ano es valido
        if ano:
            if ano == 1 and year > 2024:
                datoyear = year - 1
            else:
                datoyear = year
        else:
            datoyear = year
        #print ('datoyear: ',datoyear)

        # Se crea la funcion de Excel
        def excel(fecha):
            
            from openpyxl import load_workbook

            FILE_PATH ='C:/QPTECH/EXCEL/PLANTILLAS/ConsumoDeAguaMolino3_Plantilla.xlsx'
            sheet3 = 'Mensual'
            
            workbook= load_workbook(FILE_PATH,read_only=False)
            ws3=workbook[sheet3]
            # Colocar el año en la Fila 1 Columna A
            ws3.cell(row = 2,column = 1, value = datoyear)
            # Carga de la Columna B
            rowIndex = 2
            for filas in colu2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 2, value = filas)
                rowIndex +=1
            # Carga de la Columna C
            rowIndex = 2
            for filas in colu4:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 3, value = filas)
                rowIndex +=1
            # Carga de la Columna D
            rowIndex = 2
            for filas in colu1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 4, value = filas)
                rowIndex +=1

            fileName = f"C:/QPTECH/EXCEL/AGUA_MOLINO3/ConsumoDeAguaMolino3_{fecha[0]}-{fecha[1]}-{fecha[2]}.xlsx"
            
            workbook.save(fileName)  
            return fileName

        # Conexion a tu base de datos
        # Usa el método conectar()
        try:    
            credenciales = {
            "dbname": "agua_db",
            "user": "mantenimiento",
            "password": "#Argos2017",
            "host": "localhost",
            "port": ''
            }
            cnxn1 = psycopg2.connect(**credenciales)   
            cursor1 = cnxn1.cursor()

            #print("Iniciado PostgreSQL.....")
        
            # Determinar el ano actual
            now = datetime.now() # current date and time
            year = int(now.strftime("%Y")) # current year)
            #print('year: ',year)

            # Crear el texto del Query   
            texto_query1 = """
            SELECT year, MAX(month), MAX(fecha), MAX(volumen) FROM public."ViewAguaLanzasM3Month" 
            WHERE year = %s
            GROUP BY year
            """ % (datoyear-1) 
            # Ejecutar una consulta
            cursor1.execute(texto_query1)
            # Se crea la costante del acumulado del ano anterior
            acum = 0  
            while 1:
                row = cursor1.fetchone()
                #print(row)
                if not row:
                    break      
                #print(int(row[3]))
                acum = int(row[3])
            #print('acum: ',acum)
            # Crear el texto del Query   
            texto_query2 = """
            SELECT * FROM public."ViewAguaLanzasM3Month" 
            WHERE year = %s
            """ % (datoyear)
            # Ejecutar una consulta
            cursor1.execute(texto_query2)
            
            # Se crea la lista de los datos
            ctto = 0
            dato = 0  
            while 1:
                row = cursor1.fetchone()
                ctto = ctto + 1
                if not row:
                    break
                if ctto == 1:
                    dato = acum
                    colu0.append(int(row[0])) # year
                    if row[1] == 1:
                        colu1.append('ENERO') # month
                    elif row[1] == 2:
                        colu1.append('FEBRERO')
                    elif row[1] == 3:
                        colu1.append('MARZO')
                    elif row[1] == 4:
                        colu1.append('ABRIL')
                    elif row[1] == 5:
                        colu1.append('MAYO')
                    elif row[1] == 6:
                        colu1.append('JUNIO')
                    elif row[1] == 7:
                        colu1.append('JULIO')
                    elif row[1] == 8:
                        colu1.append('AGOSTO')
                    elif row[1] == 9:
                        colu1.append('SEPTIEMBRE')
                    elif row[1] == 10:
                        colu1.append('OCTUBRE')
                    elif row[1] == 11:
                        colu1.append('NOVIEMBRE')
                    elif row[1] == 12:
                        colu1.append('DICIEMBRE')
                    colu2.append(row[2].strftime("%d-%m-%Y")) # datetime
                    colu3.append(int(row[3])) # volumen
                    colu4.append(int(row[3]) - dato) # acumulado
                    dato = int(row[3]) 
                else:
                    colu0.append(int(row[0])) # year
                    if row[1] == 1:
                        colu1.append('ENERO') # month
                    elif row[1] == 2:
                        colu1.append('FEBRERO')
                    elif row[1] == 3:
                        colu1.append('MARZO')
                    elif row[1] == 4:
                        colu1.append('ABRIL')
                    elif row[1] == 5:
                        colu1.append('MAYO')
                    elif row[1] == 6:
                        colu1.append('JUNIO')
                    elif row[1] == 7:
                        colu1.append('JULIO')
                    elif row[1] == 8:
                        colu1.append('AGOSTO')
                    elif row[1] == 9:
                        colu1.append('SEPTIEMBRE')
                    elif row[1] == 10:
                        colu1.append('OCTUBRE')
                    elif row[1] == 11:
                        colu1.append('NOVIEMBRE')
                    elif row[1] == 12:
                        colu1.append('DICIEMBRE')
                    colu2.append(row[2].strftime("%d-%m-%Y")) # datetime
                    colu3.append(int(row[3])) # volumen
                    colu4.append(int(row[3]) - dato) # acumulado
                    dato = int(row[3])
                #print(ctto)
                #print(dato)
                #print(int(row[0]))
                #print(int(row[1]))   
                #print(int(row[3]))
                #print(row[2])
            #print(colu0,colu1,colu2,colu3,colu4)
            # Creo las listas X y Y
            for i in reversed(colu1):
                if listx == '':
                    listx = str(i)
                else:
                    listx = str(i) + ',' + listx

            for i in reversed(colu4):
                if listy == '':
                    listy = str(i)
                else:
                    listy = str(i) + ',' + listy
            #print(listx)
            #print(listy)
            # Ejecuto el Metodo de carga en Excel    
            excel(fecha)
            # Ejecuto el Metodo envio de correo y mensaje
            if email:
                print('email1: ',email)
                if email == 1:
                    # Envió de los correos con el archivo adjunto
                    mail("aalvarado@qptechllc.com", "Buenos días, este es un email de envio automático de Q&P Tech Panamá Inc. - Consumo de Agua en las Lanzas del Molino 3", file_name, file_path)
                    mail("jgarcia@qptechllc.com", "Buenos días, este es un email de envio automático de Q&P Tech Panamá Inc. - Consumo de Agua en las Lanzas del Molino 3", file_name, file_path)
                    mail("jcgutierrez@argos.co", "Buenos días, este es un email de envio automático de Q&P Tech Panamá Inc. - Consumo de Agua en las Lanzas del Molino 3", file_name, file_path)
                    mail("prac.ambienpan@argos.co", "Buenos días, este es un email de envio automático de Q&P Tech Panamá Inc. - Consumo de Agua en las Lanzas del Molino 3", file_name, file_path)                   
                    mail("miriam.villarreal@argos.co", "Buenos días, este es un email de envio automático de Q&P Tech Panamá Inc. - Consumo de Agua en las Lanzas del Molino 3", file_name, file_path)                   
                    env_email = 0
                elif email > 1:
                    env_email = 1
            else:
                env_email = 1
            print('email2: ',email)
        except Exception as e: 
            #print("Error de conexión PostgreSQL: ", e)
            cursor1.close()  
            cnxn1.close()
        finally:
        # Se cierra la conexion SQL y la conexion al Servidos
            cursor1.close()  
            cnxn1.close()
            #print("La conexión PostgreSQL está cerrada")

        messages.info(
            request, '  Por favor espere que carge la información')
        greeting = {}
        greeting['heading'] = "Gráfica Consumo Agua Lanzas de Molino3"
        greeting['pageview'] = "Sistema de agua de lanzas de Molino3"
        greeting['listaX'] = listx
        greeting['listaY'] = listy
        greeting['semana'] = week
        greeting['ano'] = datoyear
        greeting['anoD'] = datoyear - 1
        greeting['anoI'] = datoyear + 1
        greeting['emailE'] = env_email
        greeting['tituloX'] = "Consumo del agua de lanzas de Molino3 (lts)"

        return render(request, 'dashboard/graficaLM3A.html', greeting)


# ********************************************************************************************
# ********************************************************************************************
# VISTA GraficaConsumoAguaLanzasM3 SEMANAL =======================================================================
class GraficaConsumoAguaLanzasM3S(View):
    def get(self, request, ano = None, semana = None):
        # Declaro variables
        colu0 = list()
        colu1 = list()
        colu2 = list()
        colu3 = list()
        listx = ''
        listy = ''
        # Determinar el ano actual
        now = date.today() # current date and time
        year = date.today().year # current year
        month = date.today().month # current month
        day = date.today().day # current day
        week = date.today().isocalendar()[1] # current weeek from year
        # print(now)
        # print(year)
        # print(month)
        # print(day)
        # print(week)
        # Determinar si el ano es valido
        if ano:
            if ano == 1 and year > 2024:
                datoyear = year - 1
            else:
                datoyear = year
        else:
            datoyear = year
        #print (datoyear)
        # Determinar si se cargo un numero de semana valida
        if semana:
            if datoyear == 2024 and semana < 37:
                datoweek = week
            elif datoyear == 2024 and semana > 37 and semana <= week:
                datoweek = int(semana)   
            elif datoyear == year and semana <= week and semana in range(1,52):
                datoweek = int(semana)
            elif datoyear > 2024 and datoyear < year and semana in range(1,52):
                datoweek = int(semana)
            else:
                datoweek = week
        else:
            datoweek = week
        #print(datoweek)
                
        # Conexion a tu base de datos
        # Usa el método conectar()
        try:
            credenciales = {
                "dbname": "agua_db",
                "user": "mantenimiento",
                "password": "#Argos2017",
                "host": "localhost",
                "port": ''
            }
            cnxn1 = psycopg2.connect(**credenciales)   
            cursor1 = cnxn1.cursor()

            #print("Iniciado PostgreSQL.....")
            
            # Crear el texto del Query   
            texto_query1 = """
            SELECT MIN(fecha) FROM public."ViewAguaLanzasM3Day"
            WHERE year = %s AND week = %s
            """ % (datoyear, datoweek)
            # Ejecutar una consulta
            cursor1.execute(texto_query1)
            # Se busca la fecha del dia anterior a esta semana 
            while 1:
                row = cursor1.fetchone()
                #print(row) 
                if not row:
                    break      
                daytxt = row[0].strftime("%Y-%m-%d")
                #print(daytxt) # day min weed
                fechaold = datetime.strptime(daytxt,"%Y-%m-%d") - timedelta(days=1)  # day subtraction operation
                #print(fechaold)
                datofecha = fechaold.strftime("%Y-%m-%d") # old date
                #print(datofecha)
            # Crear el texto del Query   
            texto_query1 = """
            SELECT MAX(fecha), MAX(volumen) FROM public."ViewAguaLanzasM3Day"
            WHERE TO_CHAR (fecha:: DATE, 'yyyy-mm-dd') <= TO_CHAR ('/%s/':: DATE, 'yyyy-mm-dd')
            """ % (datofecha)
            # Ejecutar una consulta
            cursor1.execute(texto_query1)
            # Se carga a una variable el acumulado del dia anterior
            acum = 0 
            while 1:
                row = cursor1.fetchone()
                #print(row) 
                if not row:
                    break 
                acum = int(row[1])
                #print(int(row[1]))
            #print(acum)
            # Crear el texto del Query   
            texto_query1 = """
            SELECT fecha, volumen FROM public."ViewAguaLanzasM3Day"
            WHERE year = %s AND week = %s
            """ % (datoyear, datoweek)
            # Ejecutar una consulta
            cursor1.execute(texto_query1)
            # Se crea la lista de los datos
            ctto = 0
            dato = 0  
            while 1:
                row = cursor1.fetchone()
                #print(row)
                ctto = ctto + 1
                if not row:
                    break
                if ctto == 1:
                    dato = acum
                    colu0.append(row[0].strftime("%d-%m-%Y")) # datetime
                    colu1.append(int(row[1])) # volumen
                    colu2.append(int(row[1]) - dato) # acumulado
                    #print(row[0].weekday())
                    if row[0].weekday() == 0:
                        colu3.append('LUNES - ')
                    elif row[0].weekday() == 1:
                        colu3.append('MARTES - ')
                    elif row[0].weekday() == 2:
                        colu3.append('MIERCOLES - ')            
                    elif row[0].weekday() == 3:
                        colu3.append('JUEVES - ')
                    elif row[0].weekday() == 4:
                        colu3.append('VIERNES - ')
                    elif row[0].weekday() == 5:
                        colu3.append('SABADO - ') 
                    elif row[0].weekday() == 6:
                        colu3.append('DOMINGO - ')
                    dato = int(row[1])
                else:
                    colu0.append(row[0].strftime("%d-%m-%Y")) # datetime
                    colu1.append(int(row[1])) # volumen
                    colu2.append(int(row[1]) - dato) # acumulado
                    #print(row[0].weekday())
                    if row[0].weekday() == 0:
                        colu3.append('LUNES - ')
                    elif row[0].weekday() == 1:
                        colu3.append('MARTES - ')
                    elif row[0].weekday() == 2:
                        colu3.append('MIERCOLES - ')            
                    elif row[0].weekday() == 3:
                        colu3.append('JUEVES - ')
                    elif row[0].weekday() == 4:
                        colu3.append('VIERNES - ')
                    elif row[0].weekday() == 5:
                        colu3.append('SABADO - ') 
                    elif row[0].weekday() == 6:
                        colu3.append('DOMINGO - ')                    
                    dato = int(row[1])            
            #print(colu0,colu1,colu2,colu3)
            # Creo las listas X y Y
            for i in zip(reversed(colu0),reversed(colu3)):
                if listx == '':
                    listx = str(i[1]) + str(i[0]) 
                else:
                    listx = str(i[1]) + str(i[0]) + ',' + listx

            for i in reversed(colu2):
                if listy == '':
                    listy = str(i)
                else:
                    listy = str(i) + ',' + listy
            #print(listx)
            #print(listy)
        except Exception as e: 
            #print("Error de conexión PostgreSQL: ", e)
            cursor1.close()  
            cnxn1.close()
        finally:
        # Se cierra la conexion SQL y la conexion al Servidos
            cursor1.close()  
            cnxn1.close()
            #print("La conexión PostgreSQL está cerrada")
    
        messages.info(
            request, '  Por favor espere que carge la información')
        greeting = {}
        greeting['heading'] = "Gráfica Consumo Agua Lanzas Molino3 Semanal"
        greeting['pageview'] = "Sistema de agua de lanzas de Molino3"
        greeting['listaX'] = listx
        greeting['listaY'] = listy
        greeting['semana'] = datoweek
        greeting['semanaD'] = datoweek - 1
        greeting['semanaI'] = datoweek + 1
        greeting['ano'] = datoyear
        greeting['tituloX'] = "Consumo del agua de lanzas de Molino3 (lts)"

        return render(request, 'dashboard/graficaLM3S.html', greeting)
