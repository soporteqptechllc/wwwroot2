# Import pytz
import pytz
from itertools import groupby
# Import Django
from django.shortcuts import render
from django.db.models import Count,Sum,Max,Min
from django.views.generic import ( 
    View,
    ListView, 
    DetailView,
    CreateView,
    TemplateView,
    UpdateView,
    DeleteView
    )
# Local Models de la BD
from .models import (
    boquillas,
    productos,
    mezclado
)
# Import from datetime of Python 
from datetime import date, datetime, timedelta, timezone
# Import send_mail  
from .send_mail import mail
# Import CalculoTonSilo
from CalculoTonSilo import calculo_ton_silos
#from applications.boquillas.send_mail import mail
# Create your views here.
# ****************************************************
# Lista todos los registros por boquillas
# ****************************************************
class ListBoquillasAll(ListView):
    template_name = 'boquillas/listar_boquillas_all.html'
    paginate_by = 100 # es el numero de filas por pagina
    #ordering = '-id'# el signo - lo ordena de mayor a menor
    #model = boquillas # Se elimina si se usa el metod get
    queryset = boquillas.objects.all().order_by('-id')[:1000]
    # Se puede usar una variable de contexto en vez de usar en pantalla el "object_list"
    context_object_name = 'list_boq'

# ****************************************************    
# Lista todos los registros por una boquillas
# **************************************************** 
class ListBoquillaView(View):
    # Metodo que busca y crea las listas para ser mostrada en pantalla
    def get(self , request, kword=None):
        #print('*********GET*********')
        # ********************************************    
        # Inicio
        # ******************************************** 
        
        if kword >=1 and kword <= 3:
            # ********************************************    
            # Datos para la tabla
            # ********************************************
            list_boq = boquillas.objects.filter(n_boquilla = kword).order_by('-id')[:120]
        greeting = {}

        greeting['heading'] = "Producción Actual 1"
        greeting['pageview'] = "Producción Actual 2"
        # greeting['heading'] = "Producción Actual Diaría En Cremas"+" - Lote: "+ str(idlote)
        # greeting['pageview'] = "Fecha: "+ fecha2 +" - "+"Producto: "+ Sem
        # greeting['fechatxt'] = fecha2
        greeting['kbq'] = kword
        greeting['sub_txt1'] = "Tabla de la boquilla # "+ str(kword)
        greeting['sub_txt2'] = "Reporte de la boquilla # "+ str(kword)
        greeting['list_boq'] = list_boq
        greeting['vfech'] = 0

        return render (request,'boquillas/listar_boquillas.html',greeting)

# ****************************************************    
# Lista un resumen diario del proceso de ensacado
# **************************************************** 
class ResumenDiarioView(View):
    # Metodo que busca y crea las listas para ser mostrada en pantalla
    def get(self , request):
        #print('*********GET*********')
        # ********************************************    
        # Inicio
        # ********************************************
        now = datetime.now()  # current date and time
        # Declaracion de variables
        tz = pytz.timezone('America/Bogota') # current time zone local
        now = datetime.now() # current date and time

        # Busco el ultimo dato guardado y tomo la fecha
        list_fech =  boquillas.objects.all().order_by('-id').first()
        if list_fech:
            fech = list_fech.datetime_plc.astimezone(tz).strftime("%Y-%m-%d")
            #print('datetime_plc: ',fech)
            fech1 = datetime.strptime(fech,"%Y-%m-%d")
            #print('fech1: ',fech1)
            fech1z = tz.localize(fech1)
            #print('fech1z: ',fech1z)
            fech2 = fech1 + timedelta(days=1)  # day addition operation
            fech2z = tz.localize(fech2)
            #print('fech2z: ',fech2z)
            fecht1z = tz.localize(fech1)
            fecht4z = tz.localize(fech2)
            fecht2z = fecht1z + timedelta(hours=7)  # hours addition operation
            fecht3z = fecht1z + timedelta(hours=16,minutes=30)  # hours addition operation
            #print('fecha t1: ',fecht1z)
            #print('fecha t2: ',fecht2z)
            #print('fecha t3: ',fecht3z)
            #print('fecha t4: ',fecht4z)
            # ********************************************    
            # Datos para la tabla #1
            # ********************************************
            # Se defines las variables
            colu1 = []
            colu2 = []
            colu3 = []
            colu4 = []
            colu1t1 = []
            colu2t1 = []
            colu3t1 = []
            colu4t1 = []
            colu5t1 = []
            colu6t1 = []
            colu7t1 = []
            colu8t1 = []
            colu9t1 = []
            colu1t2 = []
            colu2t2 = []
            colu3t2 = []
            colu4t2 = []
            colu5t2 = []
            colu6t2 = []
            colu7t2 = []
            colu8t2 = []
            colu9t2 = []
            colu1t3 = []
            colu2t3 = []
            colu3t3 = []
            colu4t3 = []
            colu5t3 = []
            colu6t3 = []
            colu7t3 = []
            colu8t3 = []
            colu9t3 = []            
            # Se guarda los productos del dia
            list_ens = boquillas.objects.filter(datetime_plc__gte = fech1z, datetime_plc__lt = fech2z).order_by('id')                       

            if list_ens:
                # Se crea una lista por producto por dia y se cuenta               
                list_ensp = list_ens.values('id_producto').order_by('id_producto').annotate(count=Count('id_producto'))
                for i in list_ensp:
                    #print('id_prod: ',i['id_producto'])
                    idp = i['id_producto']
                    prod = productos.objects.get(id = idp )
                    peso = float(str(prod.peso))
                    #print(peso)
                    sum_saco = 0
                    # Busco el ultimo numero de sacos y el primer numero de sacos por boqillas
                    lsb1t1l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                    lsb1t1f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                    lsb1t2l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                    lsb1t2f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                    lsb1t3l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                    lsb1t3f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()

                    lsb2t1l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                    lsb2t1f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                    lsb2t2l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                    lsb2t2f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                    lsb2t3l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                    lsb2t3f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()

                    lsb3t1l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                    lsb3t1f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                    lsb3t2l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                    lsb3t2f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                    lsb3t3l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                    lsb3t3f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()                                       
                    # Verifico si existen datos
                    if lsb1t1l and lsb1t1f:
                        # print('datos2: ',lsb1t1l.id,lsb1t1l.idbq)
                        # print('datos1: ',lsb1t1f.id,lsb1t1f.idbq)
                        sum_saco = sum_saco + (lsb1t1l.idbq - lsb1t1f.idbq + 1)
                        #print('suma1: ',sum_saco)
                    if lsb1t2l and lsb1t2f:
                        # print('datos2: ',lsb1t2l.id,lsb1t2l.idbq)
                        # print('datos1: ',lsb1t2f.id,lsb1t2f.idbq)
                        sum_saco = sum_saco + (lsb1t2l.idbq - lsb1t2f.idbq + 1)
                        #print('suma2: ',sum_saco)
                    if lsb1t3l and lsb1t3f:
                        # print('datos2: ',lsb1t3l.id,lsb1t3l.idbq)
                        # print('datos1: ',lsb1t3f.id,lsb1t3f.idbq)
                        sum_saco = sum_saco + (lsb1t3l.idbq - lsb1t3f.idbq + 1)
                        #print('suma3: ',sum_saco)

                    if lsb2t1l and lsb2t1f:
                        # print('datos2: ',lsb2t1l.id,lsb2t1l.idbq)
                        # print('datos1: ',lsb2t1f.id,lsb2t1f.idbq)
                        sum_saco = sum_saco + (lsb2t1l.idbq - lsb2t1f.idbq + 1)
                        #print('suma4: ',sum_saco)
                    if lsb2t2l and lsb2t2f:
                        # print('datos2: ',lsb2t2l.id,lsb2t2l.idbq)
                        # print('datos1: ',lsb2t2f.id,lsb2t2f.idbq)
                        sum_saco = sum_saco + (lsb2t2l.idbq - lsb2t2f.idbq + 1)
                        #print('suma5: ',sum_saco)
                    if lsb2t3l and lsb2t3f:
                        # print('datos2: ',lsb2t3l.id,lsb2t3l.idbq)
                        # print('datos1: ',lsb2t3f.id,lsb2t3f.idbq)
                        sum_saco = sum_saco + (lsb2t3l.idbq - lsb2t3f.idbq + 1)
                        #print('suma6: ',sum_saco)

                    if lsb3t1l and lsb3t1f:
                        # print('datos2: ',lsb3t1l.id,lsb3t1l.idbq)
                        # print('datos1: ',lsb3t1f.id,lsb3t1f.idbq)
                        sum_saco = sum_saco + (lsb3t1l.idbq - lsb3t1f.idbq + 1)
                        #print('suma7: ',sum_saco)
                    if lsb3t2l and lsb3t2f:
                        # print('datos2: ',lsb3t2l.id,lsb3t2l.idbq)
                        # print('datos1: ',lsb3t2f.id,lsb3t2f.idbq)
                        sum_saco = sum_saco + (lsb3t2l.idbq - lsb3t2f.idbq + 1)
                        #print('suma8: ',sum_saco)
                    if lsb3t3l and lsb3t3f:
                        # print('datos2: ',lsb3t3l.id,lsb3t3l.idbq)
                        # print('datos1: ',lsb3t3f.id,lsb3t3f.idbq)
                        sum_saco = sum_saco + (lsb3t3l.idbq - lsb3t3f.idbq + 1)
                        #print('suma9: ',sum_saco)
                    ton_saco = round((sum_saco * peso)/1000,1)
                    if sum_saco > 1:
                        colu1.append(prod.id_Producto)
                        colu2.append(prod.name)
                        colu3.append(sum_saco)
                        colu4.append(ton_saco)

                #for i in zip(colu1,colu2,colu3,colu4):
                #    print(i)
                # Se crea un unica tabla
                resultado1 = zip(colu1,colu2,colu3,colu4)

                # ****************************************************
                # Se crean las tablas de produccion por turnos del dia  
                list_enst1 = boquillas.objects.filter(datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z).order_by('id')
                list_enst2 = boquillas.objects.filter(datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z).order_by('id')
                list_enst3 = boquillas.objects.filter(datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z).order_by('id')                      
                # Verifico si existen datos
                if list_enst1:
                    # Se crea una lista por producto por turno y se suma el PESO                
                    list_enst1p = list_enst1.values('id_producto','name_producto').order_by('id_producto').annotate(TotalPeso=Sum('peso'))

                    for i in list_enst1p:               
                        #print('id_prod: ',i['id_producto'])
                        #print('name_prod: ',i['name_producto'])
                        #print('totalkg: ',i['TotalPeso'])
                        idp = i['id_producto']
                        peso_prod = i['TotalPeso']
                        tonreal = round(peso_prod /1000,1)
                        prod = productos.objects.get(id = idp )
                        peso = float(str(prod.peso))
                        sum_saco = 0
                        # Buscar el ultimo numero de sacos y el primer numero de sacos por boqillas
                        lsb1l = list_enst1.filter(id_producto = prod,n_boquilla = 1).order_by('id').last()
                        lsb1f = list_enst1.filter(id_producto = prod,n_boquilla = 1).order_by('id').first()
                        lsb2l = list_enst1.filter(id_producto = prod,n_boquilla = 2).order_by('id').last()
                        lsb2f = list_enst1.filter(id_producto = prod,n_boquilla = 2).order_by('id').first()
                        lsb3l = list_enst1.filter(id_producto = prod,n_boquilla = 3).order_by('id').last()
                        lsb3f = list_enst1.filter(id_producto = prod,n_boquilla = 3).order_by('id').first()                                        
                        # Verifico si existen datos
                        if lsb1l and lsb1f:
                            #print('datosl: ',lsb1l.id,lsb1l.idbq)
                            #print('datosf: ',lsb1f.id,lsb1f.idbq)
                            sum_saco = lsb1l.idbq - lsb1f.idbq + 1
                            #print('suma1: ',sum_saco)
                        if lsb2l and lsb2f:
                            sum_saco = sum_saco + (lsb2l.idbq - lsb2f.idbq + 1)
                            #print('suma2: ',sum_saco)
                        if lsb3l and lsb3f:
                            sum_saco = sum_saco + (lsb3l.idbq - lsb3f.idbq + 1)
                            #print('suma3: ',sum_saco)
                        ton_saco = round((sum_saco * peso)/1000,1)
                        if sum_saco > 1:
                            colu1t1.append(prod.id_Producto)
                            colu2t1.append(prod.name)
                            colu3t1.append(sum_saco)
                            colu4t1.append(ton_saco)                                       
                            colu5t1.append(tonreal)
                        # Buscar los ultimos datos de hora y fecha y el primer dato de hora y fecha
                        list_enst1pl = list_enst1.filter(id_producto = prod).order_by('id').last()
                        list_enst1pf = list_enst1.filter(id_producto = prod).order_by('id').first()
                        # Verifico si existen datos
                        if sum_saco > 1 and list_enst1pl and list_enst1pf:
                            #print('datosL: ',list_enst1pl.hora,list_enst1pl.datetime_plc)
                            #print('datosF: ',list_enst1pf.hora,list_enst1pf.datetime_plc)
                            hor_ini = list_enst1pf.hora
                            hor_fin = list_enst1pl.hora
                            fec_ini = list_enst1pf.datetime_plc
                            fec_fin = list_enst1pl.datetime_plc
                            dif_fec = fec_fin - fec_ini
                            #print('dif_fec: ',dif_fec)
                            dif_fec_txt = str(dif_fec)
                            colu6t1.append(hor_ini)
                            colu7t1.append(hor_fin)
                            colu8t1.append(dif_fec_txt)
                            colu9t1.append('')
                        
                # for i in zip(colu1t1,colu2t1,colu3t1,colu4t1,colu5t1,colu6t1,colu7t1,colu8t1,colu9t1):
                #     print(i)
                # Se crea un unica tabla
                resultado2 = zip(colu1t1,colu2t1,colu3t1,colu4t1,colu5t1,colu6t1,colu7t1,colu8t1,colu9t1)

                # Verifico si existen datos
                if list_enst2:
                    # Se crea una lista por producto por turno y se suma el PESO                
                    list_enst2p = list_enst2.values('id_producto','name_producto').order_by('id_producto').annotate(TotalPeso=Sum('peso'))

                    for i in list_enst2p:               
                        #print('id_prod: ',i['id_producto'])
                        #print('name_prod: ',i['name_producto'])
                        #print('totalkg: ',i['TotalPeso'])
                        idp = i['id_producto']
                        peso_prod = i['TotalPeso']
                        tonreal = round(peso_prod /1000,1)
                        prod = productos.objects.get(id = idp )
                        peso = float(str(prod.peso))
                        sum_saco = 0
                        # Buscar el ultimo numero de sacos y el primer numero de sacos por boqillas
                        lsb1l = list_enst2.filter(id_producto = prod,n_boquilla = 1).order_by('id').last()
                        lsb1f = list_enst2.filter(id_producto = prod,n_boquilla = 1).order_by('id').first()
                        lsb2l = list_enst2.filter(id_producto = prod,n_boquilla = 2).order_by('id').last()
                        lsb2f = list_enst2.filter(id_producto = prod,n_boquilla = 2).order_by('id').first()
                        lsb3l = list_enst2.filter(id_producto = prod,n_boquilla = 3).order_by('id').last()
                        lsb3f = list_enst2.filter(id_producto = prod,n_boquilla = 3).order_by('id').first()                                        
                        # Verifico si existen datos
                        if lsb1l and lsb1f:
                            #print('datosl: ',lsb1l.id,lsb1l.idbq)
                            #print('datosf: ',lsb1f.id,lsb1f.idbq)
                            sum_saco = lsb1l.idbq - lsb1f.idbq + 1
                            #print('suma1: ',sum_saco)
                        if lsb2l and lsb2f:
                            sum_saco = sum_saco + (lsb2l.idbq - lsb2f.idbq + 1)
                            #print('suma2: ',sum_saco)
                        if lsb3l and lsb3f:
                            sum_saco = sum_saco + (lsb3l.idbq - lsb3f.idbq + 1)
                            #print('suma3: ',sum_saco)
                        ton_saco = round((sum_saco * peso)/1000,1)
                        if sum_saco > 1:
                            colu1t2.append(prod.id_Producto)
                            colu2t2.append(prod.name)
                            colu3t2.append(sum_saco)
                            colu4t2.append(ton_saco)                                       
                            colu5t2.append(tonreal)
                        # Buscar los ultimos datos de hora y fecha y el primer dato de hora y fecha
                        list_enst2pl = list_enst2.filter(id_producto = prod).order_by('id').last()
                        list_enst2pf = list_enst2.filter(id_producto = prod).order_by('id').first()
                        # Verifico si existen datos
                        if sum_saco > 1 and list_enst2pl and list_enst2pf:
                            #print('datosL: ',list_enst2pl.hora,list_enst2pl.datetime_plc)
                            #print('datosF: ',list_enst2pf.hora,list_enst2pf.datetime_plc)
                            hor_ini = list_enst2pf.hora
                            hor_fin = list_enst2pl.hora
                            fec_ini = list_enst2pf.datetime_plc
                            fec_fin = list_enst2pl.datetime_plc
                            dif_fec = fec_fin - fec_ini
                            #print('dif_fec: ',dif_fec)
                            dif_fec_txt = str(dif_fec)
                            colu6t2.append(hor_ini)
                            colu7t2.append(hor_fin)
                            colu8t2.append(dif_fec_txt)
                            colu9t2.append('')
                        
                # for i in zip(colu1t2,colu2t2,colu3t2,colu4t2,colu5t2,colu6t2,colu7t2,colu8t2,colu9t2):
                #     print(i)
                # Se crea un unica tabla
                resultado3 = zip(colu1t2,colu2t2,colu3t2,colu4t2,colu5t2,colu6t2,colu7t2,colu8t2,colu9t2)

                # Verifico si existen datos
                if list_enst3:
                    # Se crea una lista por producto por turno y se suma el PESO                
                    list_enst3p = list_enst3.values('id_producto','name_producto').order_by('id_producto').annotate(TotalPeso=Sum('peso'))

                    for i in list_enst3p:               
                        #print('id_prod: ',i['id_producto'])
                        #print('name_prod: ',i['name_producto'])
                        #print('totalkg: ',i['TotalPeso'])
                        idp = i['id_producto']
                        peso_prod = i['TotalPeso']
                        tonreal = round(peso_prod /1000,1)
                        prod = productos.objects.get(id = idp )
                        peso = float(str(prod.peso))
                        sum_saco = 0
                        # Buscar el ultimo numero de sacos y el primer numero de sacos por boqillas
                        lsb1l = list_enst3.filter(id_producto = prod,n_boquilla = 1).order_by('id').last()
                        lsb1f = list_enst3.filter(id_producto = prod,n_boquilla = 1).order_by('id').first()
                        lsb2l = list_enst3.filter(id_producto = prod,n_boquilla = 2).order_by('id').last()
                        lsb2f = list_enst3.filter(id_producto = prod,n_boquilla = 2).order_by('id').first()
                        lsb3l = list_enst3.filter(id_producto = prod,n_boquilla = 3).order_by('id').last()
                        lsb3f = list_enst3.filter(id_producto = prod,n_boquilla = 3).order_by('id').first()                                        
                        # Verifico si existen datos
                        if lsb1l and lsb1f:
                            #print('datosl: ',lsb1l.id,lsb1l.idbq)
                            #print('datosf: ',lsb1f.id,lsb1f.idbq)
                            sum_saco = lsb1l.idbq - lsb1f.idbq + 1
                            #print('suma1: ',sum_saco)
                        if lsb2l and lsb2f:
                            sum_saco = sum_saco + (lsb2l.idbq - lsb2f.idbq + 1)
                            #print('suma2: ',sum_saco)
                        if lsb3l and lsb3f:
                            sum_saco = sum_saco + (lsb3l.idbq - lsb3f.idbq + 1)
                            #print('suma3: ',sum_saco)
                        ton_saco = round((sum_saco * peso)/1000,1)
                        if sum_saco > 1:
                            colu1t3.append(prod.id_Producto)
                            colu2t3.append(prod.name)
                            colu3t3.append(sum_saco)
                            colu4t3.append(ton_saco)                                       
                            colu5t3.append(tonreal)
                        # Buscar los ultimos datos de hora y fecha y el primer dato de hora y fecha
                        list_enst3pl = list_enst3.filter(id_producto = prod).order_by('id').last()
                        list_enst3pf = list_enst3.filter(id_producto = prod).order_by('id').first()
                        # Verifico si existen datos
                        if sum_saco > 1 and list_enst3pl and list_enst3pf:
                            #print('datosL: ',list_enst3pl.hora,list_enst3pl.datetime_plc)
                            #print('datosF: ',list_enst3pf.hora,list_enst3pf.datetime_plc)
                            hor_ini = list_enst3pf.hora
                            hor_fin = list_enst3pl.hora
                            fec_ini = list_enst3pf.datetime_plc
                            fec_fin = list_enst3pl.datetime_plc
                            dif_fec = fec_fin - fec_ini
                            #print('dif_fec: ',dif_fec)
                            dif_fec_txt = str(dif_fec)
                            colu6t3.append(hor_ini)
                            colu7t3.append(hor_fin)
                            colu8t3.append(dif_fec_txt)
                            colu9t3.append('')
                        
                # for i in zip(colu1t3,colu2t3,colu3t3,colu4t3,colu5t3,colu6t3,colu7t3,colu8t3,colu9t3):
                #     print(i)
                # Se crea un unica tabla
                resultado4 = zip(colu1t3,colu2t3,colu3t3,colu4t3,colu5t3,colu6t3,colu7t3,colu8t3,colu9t3)

        greeting = {}

        greeting['heading'] = "PROCESO DE ENSACADO POR DIA"
        greeting['pageview'] = "FECHA : "+ fech
        greeting['fechatxt'] = fech
        greeting['no_fila1'] = len(colu1) + 1
        greeting['no_fila2'] = len(colu1t1) + 1
        greeting['no_fila3'] = len(colu1t2) + 1
        greeting['no_fila4'] = len(colu1t3) + 1
        greeting['sub_txt1'] = "PROCESO DE ENSACADO"
        greeting['sub_txt2'] = "Reporte de la boquilla # "
        greeting['table1'] = resultado1
        greeting['table2'] = resultado2
        greeting['table3'] = resultado3
        greeting['table4'] = resultado4
        greeting['vfech'] = 1

        return render (request,'boquillas/inf_ensacado.html',greeting)
    
    # Metodo que se ejecuta antes de validar y listar los datos
    def post(self, request):
        #print('*********POST*********')
        # ********************************************    
        # Inicio
        # ********************************************        
        now = datetime.now()  # current date and time
        # Declaracion de variables
        tz = pytz.timezone('America/Bogota') # current time zone local
        # Se lee la variable ¨fecha¨ desde el POST
        fecha = request.POST.get('fecha')
        if fecha:
            dato = datetime.strptime(fecha,"%Y%m%d")
            fech = dato.astimezone(tz).strftime("%Y-%m-%d")
        else:
            fech = now.strftime("%Y-%m-%d")
        #print("Fecha1 input: ",fecha)
        fech1 = datetime.strptime(fech,"%Y-%m-%d")
        #print('fech1: ',fech1)
        fech1z = tz.localize(fech1)
        #print('fech1z: ',fech1z)
        fech2 = fech1 + timedelta(days=1)  # day addition operation
        fech2z = tz.localize(fech2)
        #print('fech2z: ',fech2z)        
        fecht1z = tz.localize(fech1)
        fecht4z = tz.localize(fech2)
        fecht2z = fecht1z + timedelta(hours=7)  # hours addition operation
        fecht3z = fecht1z + timedelta(hours=16,minutes=30)  # hours addition operation
        # print('fecha t1: ',fecht1z)
        # print('fecha t2: ',fecht2z)
        # print('fecha t3: ',fecht3z)
        # print('fecha t4: ',fecht4z)
        # ********************************************    
        # Datos para la tabla #1
        # ********************************************
        # Se defines las variables
        colu1 = []
        colu2 = []
        colu3 = []
        colu4 = []
        colu1t1 = []
        colu2t1 = []
        colu3t1 = []
        colu4t1 = []
        colu5t1 = []
        colu6t1 = []
        colu7t1 = []
        colu8t1 = []
        colu9t1 = []
        colu1t2 = []
        colu2t2 = []
        colu3t2 = []
        colu4t2 = []
        colu5t2 = []
        colu6t2 = []
        colu7t2 = []
        colu8t2 = []
        colu9t2 = []
        colu1t3 = []
        colu2t3 = []
        colu3t3 = []
        colu4t3 = []
        colu5t3 = []
        colu6t3 = []
        colu7t3 = []
        colu8t3 = []
        colu9t3 = [] 

        # Se guarda los productos del dia
        list_ens = boquillas.objects.filter(datetime_plc__gte = fech1z, datetime_plc__lt = fech2z).order_by('id')

        if list_ens:
            # Se crea una lista por producto por dia y se cuenta               
            list_ensp = list_ens.values('id_producto').order_by('id_producto').annotate(count=Count('id_producto'))
            for i in list_ensp:
                #print('id_prod: ',i['id_producto'])
                idp = i['id_producto']
                prod = productos.objects.get(id = idp )
                peso = float(str(prod.peso))
                #print(peso)
                sum_saco = 0
                # Busco el ultimo numero de sacos y el primer numero de sacos por boqillas
                lsb1t1l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                lsb1t1f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                lsb1t2l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                lsb1t2f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                lsb1t3l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                lsb1t3f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()

                lsb2t1l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                lsb2t1f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                lsb2t2l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                lsb2t2f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                lsb2t3l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                lsb2t3f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()

                lsb3t1l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                lsb3t1f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                lsb3t2l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                lsb3t2f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                lsb3t3l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                lsb3t3f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()                                       
                # Verifico si existen datos
                if lsb1t1l and lsb1t1f:
                    # print('datos2: ',lsb1t1l.id,lsb1t1l.idbq)
                    # print('datos1: ',lsb1t1f.id,lsb1t1f.idbq)
                    sum_saco = sum_saco + (lsb1t1l.idbq - lsb1t1f.idbq + 1)
                    #print('suma1: ',sum_saco)
                if lsb1t2l and lsb1t2f:
                    # print('datos2: ',lsb1t2l.id,lsb1t2l.idbq)
                    # print('datos1: ',lsb1t2f.id,lsb1t2f.idbq)
                    sum_saco = sum_saco + (lsb1t2l.idbq - lsb1t2f.idbq + 1)
                    #print('suma2: ',sum_saco)
                if lsb1t3l and lsb1t3f:
                    # print('datos2: ',lsb1t3l.id,lsb1t3l.idbq)
                    # print('datos1: ',lsb1t3f.id,lsb1t3f.idbq)
                    sum_saco = sum_saco + (lsb1t3l.idbq - lsb1t3f.idbq + 1)
                    #print('suma3: ',sum_saco)

                if lsb2t1l and lsb2t1f:
                    # print('datos2: ',lsb2t1l.id,lsb2t1l.idbq)
                    # print('datos1: ',lsb2t1f.id,lsb2t1f.idbq)
                    sum_saco = sum_saco + (lsb2t1l.idbq - lsb2t1f.idbq + 1)
                    #print('suma4: ',sum_saco)
                if lsb2t2l and lsb2t2f:
                    # print('datos2: ',lsb2t2l.id,lsb2t2l.idbq)
                    # print('datos1: ',lsb2t2f.id,lsb2t2f.idbq)
                    sum_saco = sum_saco + (lsb2t2l.idbq - lsb2t2f.idbq + 1)
                    #print('suma5: ',sum_saco)
                if lsb2t3l and lsb2t3f:
                    # print('datos2: ',lsb2t3l.id,lsb2t3l.idbq)
                    # print('datos1: ',lsb2t3f.id,lsb2t3f.idbq)
                    sum_saco = sum_saco + (lsb2t3l.idbq - lsb2t3f.idbq + 1)
                    #print('suma6: ',sum_saco)

                if lsb3t1l and lsb3t1f:
                    # print('datos2: ',lsb3t1l.id,lsb3t1l.idbq)
                    # print('datos1: ',lsb3t1f.id,lsb3t1f.idbq)
                    sum_saco = sum_saco + (lsb3t1l.idbq - lsb3t1f.idbq + 1)
                    #print('suma7: ',sum_saco)
                if lsb3t2l and lsb3t2f:
                    # print('datos2: ',lsb3t2l.id,lsb3t2l.idbq)
                    # print('datos1: ',lsb3t2f.id,lsb3t2f.idbq)
                    sum_saco = sum_saco + (lsb3t2l.idbq - lsb3t2f.idbq + 1)
                    #print('suma8: ',sum_saco)
                if lsb3t3l and lsb3t3f:
                    # print('datos2: ',lsb3t3l.id,lsb3t3l.idbq)
                    # print('datos1: ',lsb3t3f.id,lsb3t3f.idbq)
                    sum_saco = sum_saco + (lsb3t3l.idbq - lsb3t3f.idbq + 1)
                    #print('suma9: ',sum_saco)                
                ton_saco = round((sum_saco * peso)/1000,1)
                if sum_saco > 1:
                    colu1.append(prod.id_Producto)
                    colu2.append(prod.name)
                    colu3.append(sum_saco)
                    colu4.append(ton_saco)

        #for i in zip(colu1,colu2,colu3,colu4):
        #    print(i)
        # Se crea un unica tabla
        resultado1 = zip(colu1,colu2,colu3,colu4)
        
        # ****************************************************
        # Se crean las tablas de produccion por turnos del dia  
        list_enst1 = boquillas.objects.filter(datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z).order_by('id')
        list_enst2 = boquillas.objects.filter(datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z).order_by('id')
        list_enst3 = boquillas.objects.filter(datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z).order_by('id')                      
        # Verifico si existen datos
        if list_enst1:
            # Se crea una lista por producto por turno y se suma el PESO                
            list_enst1p = list_enst1.values('id_producto','name_producto').order_by('id_producto').annotate(TotalPeso=Sum('peso'))

            for i in list_enst1p:               
                #print('id_prod: ',i['id_producto'])
                #print('name_prod: ',i['name_producto'])
                #print('totalkg: ',i['TotalPeso'])
                idp = i['id_producto']
                peso_prod = i['TotalPeso']
                tonreal = round(peso_prod /1000,1)
                prod = productos.objects.get(id = idp )
                peso = float(str(prod.peso))
                sum_saco = 0
                # Buscar el ultimo numero de sacos y el primer numero de sacos por boqillas
                lsb1l = list_enst1.filter(id_producto = prod,n_boquilla = 1).order_by('id').last()
                lsb1f = list_enst1.filter(id_producto = prod,n_boquilla = 1).order_by('id').first()
                lsb2l = list_enst1.filter(id_producto = prod,n_boquilla = 2).order_by('id').last()
                lsb2f = list_enst1.filter(id_producto = prod,n_boquilla = 2).order_by('id').first()
                lsb3l = list_enst1.filter(id_producto = prod,n_boquilla = 3).order_by('id').last()
                lsb3f = list_enst1.filter(id_producto = prod,n_boquilla = 3).order_by('id').first()                                        
                # Verifico si existen datos
                if lsb1l and lsb1f:
                    #print('datosl: ',lsb1l.id,lsb1l.idbq)
                    #print('datosf: ',lsb1f.id,lsb1f.idbq)
                    sum_saco = lsb1l.idbq - lsb1f.idbq + 1
                    #print('suma1: ',sum_saco)
                if lsb2l and lsb2f:
                    #print('datosl: ',lsb2l.id,lsb2l.idbq)
                    #print('datosf: ',lsb2f.id,lsb2f.idbq)
                    sum_saco = sum_saco + (lsb2l.idbq - lsb2f.idbq + 1)
                    #print('suma2: ',sum_saco)
                if lsb3l and lsb3f:
                    sum_saco = sum_saco + (lsb3l.idbq - lsb3f.idbq + 1)
                    #print('suma3: ',sum_saco)
                ton_saco = round((sum_saco * peso)/1000,1)
                if sum_saco > 1:
                    colu1t1.append(prod.id_Producto)
                    colu2t1.append(prod.name)
                    colu3t1.append(sum_saco)
                    colu4t1.append(ton_saco)                                       
                    colu5t1.append(tonreal)
                # Buscar los ultimos datos de hora y fecha y el primer dato de hora y fecha
                list_enst1pl = list_enst1.filter(id_producto = prod).order_by('id').last()
                list_enst1pf = list_enst1.filter(id_producto = prod).order_by('id').first()
                # Verifico si existen datos
                if sum_saco > 1 and list_enst1pl and list_enst1pf:
                    #print('datosL: ',list_enst1pl.hora,list_enst1pl.datetime_plc)
                    #print('datosF: ',list_enst1pf.hora,list_enst1pf.datetime_plc)
                    hor_ini = list_enst1pf.hora
                    hor_fin = list_enst1pl.hora
                    fec_ini = list_enst1pf.datetime_plc
                    fec_fin = list_enst1pl.datetime_plc
                    dif_fec = fec_fin - fec_ini
                    #print('dif_fec: ',dif_fec)
                    dif_fec_txt = str(dif_fec)
                    colu6t1.append(hor_ini)
                    colu7t1.append(hor_fin)
                    colu8t1.append(dif_fec_txt)
                    colu9t1.append('')
                
        # for i in zip(colu1t1,colu2t1,colu3t1,colu4t1,colu5t1,colu6t1,colu7t1,colu8t1,colu9t1):
        #     print(i)
        # Se crea un unica tabla
        resultado2 = zip(colu1t1,colu2t1,colu3t1,colu4t1,colu5t1,colu6t1,colu7t1,colu8t1,colu9t1)

        # Verifico si existen datos
        if list_enst2:
            # Se crea una lista por producto por turno y se suma el PESO                
            list_enst2p = list_enst2.values('id_producto','name_producto').order_by('id_producto').annotate(TotalPeso=Sum('peso'))

            for i in list_enst2p:               
                #print('id_prod: ',i['id_producto'])
                #print('name_prod: ',i['name_producto'])
                #print('totalkg: ',i['TotalPeso'])
                idp = i['id_producto']
                peso_prod = i['TotalPeso']
                tonreal = round(peso_prod /1000,1)
                prod = productos.objects.get(id = idp )
                peso = float(str(prod.peso))
                sum_saco = 0
                # Buscar el ultimo numero de sacos y el primer numero de sacos por boqillas
                lsb1l = list_enst2.filter(id_producto = prod,n_boquilla = 1).order_by('id').last()
                lsb1f = list_enst2.filter(id_producto = prod,n_boquilla = 1).order_by('id').first()
                lsb2l = list_enst2.filter(id_producto = prod,n_boquilla = 2).order_by('id').last()
                lsb2f = list_enst2.filter(id_producto = prod,n_boquilla = 2).order_by('id').first()
                lsb3l = list_enst2.filter(id_producto = prod,n_boquilla = 3).order_by('id').last()
                lsb3f = list_enst2.filter(id_producto = prod,n_boquilla = 3).order_by('id').first()                                        
                # Verifico si existen datos
                if lsb1l and lsb1f:
                    #print('datosl: ',lsb1l.id,lsb1l.idbq)
                    #print('datosf: ',lsb1f.id,lsb1f.idbq)
                    sum_saco = lsb1l.idbq - lsb1f.idbq + 1
                    #print('suma1: ',sum_saco)
                if lsb2l and lsb2f:
                    #print('datosl: ',lsb2l.id,lsb2l.idbq)
                    #print('datosf: ',lsb2f.id,lsb2f.idbq)
                    sum_saco = sum_saco + (lsb2l.idbq - lsb2f.idbq + 1)
                    #print('suma2: ',sum_saco)
                if lsb3l and lsb3f:
                    sum_saco = sum_saco + (lsb3l.idbq - lsb3f.idbq + 1)
                    #print('suma3: ',sum_saco)
                ton_saco = round((sum_saco * peso)/1000,1)
                if sum_saco > 1:
                    colu1t2.append(prod.id_Producto)
                    colu2t2.append(prod.name)
                    colu3t2.append(sum_saco)
                    colu4t2.append(ton_saco)                                       
                    colu5t2.append(tonreal)
                # Buscar los ultimos datos de hora y fecha y el primer dato de hora y fecha
                list_enst2pl = list_enst2.filter(id_producto = prod).order_by('id').last()
                list_enst2pf = list_enst2.filter(id_producto = prod).order_by('id').first()
                # Verifico si existen datos
                if sum_saco > 1 and list_enst2pl and list_enst2pf:
                    #print('datosL: ',list_enst2pl.hora,list_enst2pl.datetime_plc)
                    #print('datosF: ',list_enst2pf.hora,list_enst2pf.datetime_plc)
                    hor_ini = list_enst2pf.hora
                    hor_fin = list_enst2pl.hora
                    fec_ini = list_enst2pf.datetime_plc
                    fec_fin = list_enst2pl.datetime_plc
                    dif_fec = fec_fin - fec_ini
                    #print('dif_fec: ',dif_fec)
                    dif_fec_txt = str(dif_fec)
                    colu6t2.append(hor_ini)
                    colu7t2.append(hor_fin)
                    colu8t2.append(dif_fec_txt)
                    colu9t2.append('')
                
        # for i in zip(colu1t2,colu2t2,colu3t2,colu4t2,colu5t2,colu6t2,colu7t2,colu8t2,colu9t2):
        #     print(i)
        # Se crea un unica tabla
        resultado3 = zip(colu1t2,colu2t2,colu3t2,colu4t2,colu5t2,colu6t2,colu7t2,colu8t2,colu9t2)

        # Verifico si existen datos
        if list_enst3:
            # Se crea una lista por producto por turno y se suma el PESO                
            list_enst3p = list_enst3.values('id_producto','name_producto').order_by('id_producto').annotate(TotalPeso=Sum('peso'))

            for i in list_enst3p:               
                #print('id_prod: ',i['id_producto'])
                #print('name_prod: ',i['name_producto'])
                #print('totalkg: ',i['TotalPeso'])
                idp = i['id_producto']
                peso_prod = i['TotalPeso']
                tonreal = round(peso_prod /1000,1)
                prod = productos.objects.get(id = idp )
                peso = float(str(prod.peso))
                sum_saco = 0
                # Buscar el ultimo numero de sacos y el primer numero de sacos por boqillas
                lsb1l = list_enst3.filter(id_producto = prod,n_boquilla = 1).order_by('id').last()
                lsb1f = list_enst3.filter(id_producto = prod,n_boquilla = 1).order_by('id').first()
                lsb2l = list_enst3.filter(id_producto = prod,n_boquilla = 2).order_by('id').last()
                lsb2f = list_enst3.filter(id_producto = prod,n_boquilla = 2).order_by('id').first()
                lsb3l = list_enst3.filter(id_producto = prod,n_boquilla = 3).order_by('id').last()
                lsb3f = list_enst3.filter(id_producto = prod,n_boquilla = 3).order_by('id').first()                                        
                # Verifico si existen datos
                if lsb1l and lsb1f:
                    #print('datosl: ',lsb1l.id,lsb1l.idbq)
                    #print('datosf: ',lsb1f.id,lsb1f.idbq)
                    sum_saco = lsb1l.idbq - lsb1f.idbq + 1
                    #print('suma1: ',sum_saco)
                if lsb2l and lsb2f:
                    #print('datosl: ',lsb2l.id,lsb2l.idbq)
                    #print('datosf: ',lsb2f.id,lsb2f.idbq)                    
                    sum_saco = sum_saco + (lsb2l.idbq - lsb2f.idbq + 1)
                    #print('suma2: ',sum_saco)
                if lsb3l and lsb3f:
                    sum_saco = sum_saco + (lsb3l.idbq - lsb3f.idbq + 1)
                    #print('suma3: ',sum_saco)
                ton_saco = round((sum_saco * peso)/1000,1)
                if sum_saco > 1:
                    colu1t3.append(prod.id_Producto)
                    colu2t3.append(prod.name)
                    colu3t3.append(sum_saco)
                    colu4t3.append(ton_saco)                                       
                    colu5t3.append(tonreal)
                # Buscar los ultimos datos de hora y fecha y el primer dato de hora y fecha
                list_enst3pl = list_enst3.filter(id_producto = prod).order_by('id').last()
                list_enst3pf = list_enst3.filter(id_producto = prod).order_by('id').first()
                # Verifico si existen datos
                if sum_saco > 1 and list_enst3pl and list_enst3pf:
                    #print('datosL: ',list_enst3pl.hora,list_enst3pl.datetime_plc)
                    #print('datosF: ',list_enst3pf.hora,list_enst3pf.datetime_plc)
                    hor_ini = list_enst3pf.hora
                    hor_fin = list_enst3pl.hora
                    fec_ini = list_enst3pf.datetime_plc
                    fec_fin = list_enst3pl.datetime_plc
                    dif_fec = fec_fin - fec_ini
                    #print('dif_fec: ',dif_fec)
                    dif_fec_txt = str(dif_fec)
                    colu6t3.append(hor_ini)
                    colu7t3.append(hor_fin)
                    colu8t3.append(dif_fec_txt)
                    colu9t3.append('')
                
        # for i in zip(colu1t3,colu2t3,colu3t3,colu4t3,colu5t3,colu6t3,colu7t3,colu8t3,colu9t3):
        #     print(i)
        # Se crea un unica tabla
        resultado4 = zip(colu1t3,colu2t3,colu3t3,colu4t3,colu5t3,colu6t3,colu7t3,colu8t3,colu9t3)

        greeting = {}

        greeting['heading'] = "PROCESO DE ENSACADO POR DIA"
        greeting['pageview'] = "FECHA : "+ fech
        greeting['fechatxt'] = fech
        greeting['no_fila1'] = len(colu1) + 1
        greeting['no_fila2'] = len(colu1t1) + 1
        greeting['no_fila3'] = len(colu1t2) + 1
        greeting['no_fila4'] = len(colu1t3) + 1
        greeting['sub_txt1'] = "PROCESO DE ENSACADO"
        greeting['sub_txt2'] = "Reporte de la boquilla # "
        greeting['table1'] = resultado1
        greeting['table2'] = resultado2
        greeting['table3'] = resultado3
        greeting['table4'] = resultado4
        greeting['vfech'] = 1

        return render (request,'boquillas/inf_ensacado.html',greeting)

# ****************************************************    
# Lista un resumen diario del proceso de mezclado
# **************************************************** 
class ResumenDiarioMezcladoView(View):
    # Metodo que busca y crea las listas para ser mostrada en pantalla
    def get(self , request):
        #print('*********GET*********')
        # ********************************************    
        # Inicio
        # ********************************************
        now = datetime.now()  # current date and time
        # Declaracion de variables
        tz = pytz.timezone('America/Bogota') # current time zone local

        # Busco el ultimo dato guardado y tomo la fecha
        list_fech =  mezclado.objects.all().order_by('-id').first()
        #print(list_fech)
        if list_fech:
            fech = list_fech.datetime_plc.astimezone(tz).strftime("%Y-%m-%d")
            #print('datetime_plc: ',fech)
            fech1 = datetime.strptime(fech,"%Y-%m-%d")
            #print('fech1: ',fech1)
            fech1z = tz.localize(fech1)
            #print('fech1z: ',fech1z)
            fech2 = fech1 + timedelta(days=1)  # day addition operation
            fech2z = tz.localize(fech2)
            #print('fech2z: ',fech2z)
            fecht1z = tz.localize(fech1)
            fecht4z = tz.localize(fech2)
            fecht2z = fecht1z + timedelta(hours=7)  # hours addition operation
            fecht3z = fecht1z + timedelta(hours=16,minutes=30)  # hours addition operation
            #print('fecha t1: ',fecht1z)
            #print('fecha t2: ',fecht2z)
            #print('fecha t3: ',fecht3z)
            #print('fecha t4: ',fecht4z)
            # ********************************************    
            # Datos para la tabla #1
            # ********************************************
            # Se defines las variables
            colu1 = []
            colu2 = []
            colu3 = []
            colu4 = []
            colu5 = []
            colu6 = []
            colu7 = []
            colu8 = []
            colu9 = []
            colu10 = []
            colu11 = []
            colu12 = []
            colu13 = []
            colu14 = []
            colu15 = []
            colu16 = []
            colu17 = []
            colu18 = []
            colu1t1 = []
            colu2t1 = []
            colu3t1 = []
            colu4t1 = []
            colu5t1 = []
            colu6t1 = []
            colu7t1 = []
            colu8t1 = []
            colu9t1 = []
            colu10t1 = []
            colu11t1 = []
            colu12t1 = []
            colu13t1 = []
            colu14t1 = []
            colu15t1 = []
            colu16t1 = []
            colu17t1 = []
            colu18t1 = []
            colu1t2 = []
            colu2t2 = []
            colu3t2 = []
            colu4t2 = []
            colu5t2 = []
            colu6t2 = []
            colu7t2 = []
            colu8t2 = []
            colu9t2 = []
            colu10t2 = []
            colu11t2 = []
            colu12t2 = []
            colu13t2 = []
            colu14t2 = []
            colu15t2 = []
            colu16t2 = []
            colu17t2 = []
            colu18t2 = []
            colu1t3 = []
            colu2t3 = []
            colu3t3 = []
            colu4t3 = []
            colu5t3 = []
            colu6t3 = []
            colu7t3 = []
            colu8t3 = []
            colu9t3 = []
            colu10t3 = []
            colu11t3 = []
            colu12t3 = []
            colu13t3 = []
            colu14t3 = []
            colu15t3 = []
            colu16t3 = []
            colu17t3 = []
            colu18t3 = []
            silo1t1 = []
            silo2t1 = []
            silo3t1 = []
            silo4t1 = []
            silo5t1 = []
            silo6t1 = []
            silo1t2 = []
            silo2t2 = []
            silo3t2 = []
            silo4t2 = []
            silo5t2 = []
            silo6t2 = []
            sumt1 = []
            sumt2 = []
            sumt3 = []
            sumt4 = []
            sumt5 = []
            sumt6 = []
            sumt7 = []
            sumt8 = []
            sumt9 = []
            sumt10 = []
            sumt11 = []
            sumt12 = []                        
                                                                      
            # Se guarda los productos del dia
            list_mezc = mezclado.objects.filter(datetime_plc__gte = fech1z, datetime_plc__lt = fech2z).order_by('id') 
            #for i in list_mezc:
            #   print(i.id,i.batch_id,i.lote,i.id_producto,i.name_producto,i.datetime_plc,i.bz_agre1_sp,i.bz_agre2_sp,i.bz_agre3_sp,i.bz_agre4_sp,i.bz_cem1_sp,i.bz_cem2_sp)
            # Valido si existe data
            if list_mezc:
                # Se crea una lista por producto por dia y se cuenta los SP y los PV
                list_mezcp = list_mezc.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
                fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
                sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
                sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
                sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
                list_mezcp = list_mezcp.order_by('lote')
                for i in list_mezcp:
                    #print(i)
                    #print('id_prod: ',i['id_producto'])
                    id = i['id_producto']
                    idp = productos.objects.get(id = id)
                    prod = i['name_producto']
                    lot = i['lote']
                    sum1 = round(i['sum1'],3)
                    sum2 = round(i['sum2'],3)
                    sum3 = round(i['sum3'],3)
                    sum4 = round(i['sum4'],3)
                    sum5 = round(i['sum5'],3)
                    sum6 = round(i['sum6'],3)
                    sum7 = round(i['sum7'],3)
                    sum8 = round(i['sum8'],3)
                    sum9 = round(i['sum9'],3)
                    sum10 = round(i['sum10'],3)
                    sum11 = round(i['sum11'],3)
                    sum12 = round(i['sum12'],3)                                                            
                    fecfin =i['fmx']
                    fecini =i['fmi']
                    diffec = fecfin - fecini
                    diffectxt = str(diffec)
                    #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                    # Crear las listas individuales para luego crear la general
                    colu1.append(idp)
                    colu2.append(prod)
                    colu3.append(lot)
                    colu4.append(sum1)
                    colu5.append(sum2)
                    colu6.append(sum3)
                    colu7.append(sum4)
                    colu8.append(sum5)
                    colu9.append(sum6)
                    colu10.append(sum7)
                    colu11.append(sum8)
                    colu12.append(sum9)
                    colu13.append(sum10)
                    colu14.append(sum11)
                    colu15.append(sum12)
                    colu16.append(fecfin)
                    colu17.append(fecini)
                    colu18.append(diffectxt)
                
                # Se crea una lista por year,mes y dia
                list_sumt = list_mezc.values('year_start','month_start','day_start').order_by('day_start').annotate(idf=Max('id'),
                su1=Sum('bz_agre1_sp'),su2=Sum('bz_agre1_pv'),su3=Sum('bz_agre2_sp'),su4=Sum('bz_agre2_pv'),
                su5=Sum('bz_agre3_sp'),su6=Sum('bz_agre3_pv'),su7=Sum('bz_agre4_sp'),su8=Sum('bz_agre4_pv'),
                su9=Sum('bz_cem1_sp'),su10=Sum('bz_cem1_pv'),su11=Sum('bz_cem2_sp'),su12=Sum('bz_cem2_pv'))
                # Verifico si existen datos
                if list_sumt:
                    for i in list_sumt:
                        #print(i['idf'])
                        idf = i['idf']                       
                        mescf = mezclado.objects.get(id = idf)
                        #print('silos: ',mescf.silo1,mescf.silo2,mescf.silo3,mescf.silo4,mescf.silo5,mescf.silo6)
                        sumt1.append(round(i['su1'],3))
                        sumt2.append(round(i['su2'],3))
                        sumt3.append(round(i['su3'],3))
                        sumt4.append(round(i['su4'],3))
                        sumt5.append(round(i['su5'],3))
                        sumt6.append(round(i['su6'],3))
                        sumt7.append(round(i['su7'],3))
                        sumt8.append(round(i['su8'],3))
                        sumt9.append(round(i['su9'],3))
                        sumt10.append(round(i['su10'],3))
                        sumt11.append(round(i['su11'],3))
                        sumt12.append(round(i['su12'],3))
                        silo1t1.append(round(mescf.silo1,3))
                        silo1t2.append(0)
                        Nivel_metros_s1 = round(mescf.silo1,3)
                        silo2t1.append(round(mescf.silo2,3))
                        silo2t2.append(0)
                        Nivel_metros_s2 = round(mescf.silo2,3)
                        silo3t1.append(round(mescf.silo3,3))
                        silo3t2.append(0)
                        Nivel_metros_s3 = round(mescf.silo3,3)
                        silo4t1.append(round(mescf.silo4,3))
                        silo4t2.append(0)
                        Nivel_metros_s4 = round(mescf.silo4,3)
                        silo5t1.append(round(mescf.silo5,3))
                        silo5t2.append(0)
                        Nivel_metros_s5 = round(mescf.silo5,3)
                        silo6t1.append(round(mescf.silo6,3))
                        silo6t2.append(0)
                        Nivel_metros_s6 = round(mescf.silo6,3)
            # Se crea un unica tabla
            resultado1 = zip(colu1,colu2,colu3,colu4,colu5,colu6,colu7,colu8,colu9,colu10,colu11,colu12,colu13,colu14,colu15,colu16,colu17,colu18)           
            Lista_Silos_toneladas = calculo_ton_silos(Nivel_metros_s1,Nivel_metros_s2,Nivel_metros_s3,Nivel_metros_s4,Nivel_metros_s5,Nivel_metros_s6)
            #print(Lista_Silos_toneladas)
            if Lista_Silos_toneladas:
                silo1t2[0] = Lista_Silos_toneladas[0]
                silo2t2[0] = Lista_Silos_toneladas[1]
                silo3t2[0] = Lista_Silos_toneladas[2]
                silo4t2[0] = Lista_Silos_toneladas[3]
                silo5t2[0] = Lista_Silos_toneladas[4]
                silo6t2[0] = Lista_Silos_toneladas[5]
            resultado_silo = zip(silo1t1,silo1t2,silo2t1,silo2t2,silo3t1,silo3t2,silo4t1,silo4t2,silo5t1,silo5t2,silo6t1,silo6t2)  
            resultado_suma = zip(sumt1,sumt2,sumt3,sumt4,sumt5,sumt6,sumt7,sumt8,sumt9,sumt10,sumt11,sumt12)            
            # ****************************************************
            # Se crean las tablas de produccion por turnos del dia
            list_mezct1 = mezclado.objects.filter(datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z).order_by('id')   
            list_mezct2 = mezclado.objects.filter(datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z).order_by('id')
            list_mezct3 = mezclado.objects.filter(datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z).order_by('id')  
            # Verifico si existen datos
            if list_mezct1:
                # Se crea una lista por producto por dia y se cuenta los SP y los PV               
                list_mezct1p = list_mezct1.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
                fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
                sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
                sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
                sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
                list_mezct1p = list_mezct1p.order_by('lote')
                for i in list_mezct1p:
                    #print(i)
                    #print('id_prod: ',i['id_producto'])
                    id = i['id_producto']
                    idp = productos.objects.get(id = id)
                    prod = i['name_producto']
                    lot = i['lote']
                    sum1 = round(i['sum1'],3)
                    sum2 = round(i['sum2'],3)
                    sum3 = round(i['sum3'],3)
                    sum4 = round(i['sum4'],3)
                    sum5 = round(i['sum5'],3)
                    sum6 = round(i['sum6'],3)
                    sum7 = round(i['sum7'],3)
                    sum8 = round(i['sum8'],3)
                    sum9 = round(i['sum9'],3)
                    sum10 = round(i['sum10'],3)
                    sum11 = round(i['sum11'],3)
                    sum12 = round(i['sum12'],3)
                    fecfin =i['fmx']
                    fecini =i['fmi']
                    diffec = fecfin - fecini
                    diffectxt = str(diffec)
                    #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                    # Crear las listas individuales para luego crear la general
                    colu1t1.append(idp)
                    colu2t1.append(prod)
                    colu3t1.append(lot)
                    colu4t1.append(sum1)
                    colu5t1.append(sum2)
                    colu6t1.append(sum3)
                    colu7t1.append(sum4)
                    colu8t1.append(sum5)
                    colu9t1.append(sum6)
                    colu10t1.append(sum7)
                    colu11t1.append(sum8)
                    colu12t1.append(sum9)
                    colu13t1.append(sum10)
                    colu14t1.append(sum11)
                    colu15t1.append(sum12)
                    colu16t1.append(fecfin)
                    colu17t1.append(fecini)
                    colu18t1.append(diffectxt)                   
            # Se crea un unica tabla
            resultado2 = zip(colu1t1,colu2t1,colu3t1,colu4t1,colu5t1,colu6t1,colu7t1,colu8t1,colu9t1,colu10t1,colu11t1,colu12t1,colu13t1,colu14t1,colu15t1,colu16t1,colu17t1,colu18t1)

            # Verifico si existen datos
            if list_mezct2:
                # Se crea una lista por producto por dia y se cuenta los SP y los PV               
                list_mezct2p = list_mezct2.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
                fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
                sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
                sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
                sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
                list_mezct2p = list_mezct2p.order_by('lote')
                for i in list_mezct2p:
                    #print(i)
                    #print('id_prod: ',i['id_producto'])
                    id = i['id_producto']
                    idp = productos.objects.get(id = id)
                    prod = i['name_producto']
                    lot = i['lote']
                    sum1 = round(i['sum1'],3)
                    sum2 = round(i['sum2'],3)
                    sum3 = round(i['sum3'],3)
                    sum4 = round(i['sum4'],3)
                    sum5 = round(i['sum5'],3)
                    sum6 = round(i['sum6'],3)
                    sum7 = round(i['sum7'],3)
                    sum8 = round(i['sum8'],3)
                    sum9 = round(i['sum9'],3)
                    sum10 = round(i['sum10'],3)
                    sum11 = round(i['sum11'],3)
                    sum12 = round(i['sum12'],3)
                    fecfin =i['fmx']
                    fecini =i['fmi']
                    diffec = fecfin - fecini
                    diffectxt = str(diffec)
                    #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                    # Crear las listas individuales para luego crear la general
                    colu1t2.append(idp)
                    colu2t2.append(prod)
                    colu3t2.append(lot)
                    colu4t2.append(sum1)
                    colu5t2.append(sum2)
                    colu6t2.append(sum3)
                    colu7t2.append(sum4)
                    colu8t2.append(sum5)
                    colu9t2.append(sum6)
                    colu10t2.append(sum7)
                    colu11t2.append(sum8)
                    colu12t2.append(sum9)
                    colu13t2.append(sum10)
                    colu14t2.append(sum11)
                    colu15t2.append(sum12)
                    colu16t2.append(fecfin)
                    colu17t2.append(fecini)
                    colu18t2.append(diffectxt)                   
            # Se crea un unica tabla
            resultado3 = zip(colu1t2,colu2t2,colu3t2,colu4t2,colu5t2,colu6t2,colu7t2,colu8t2,colu9t2,colu10t2,colu11t2,colu12t2,colu13t2,colu14t2,colu15t2,colu16t2,colu17t2,colu18t2)

            # Verifico si existen datos
            if list_mezct3:
                # Se crea una lista por producto por dia y se cuenta los SP y los PV               
                list_mezct3p = list_mezct3.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
                fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
                sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
                sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
                sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
                list_mezct3p = list_mezct3p.order_by('lote')
                for i in list_mezct3p:
                    #print(i)
                    #print('id_prod: ',i['id_producto'])
                    id = i['id_producto']
                    idp = productos.objects.get(id = id)
                    prod = i['name_producto']
                    lot = i['lote']
                    sum1 = round(i['sum1'],3)
                    sum2 = round(i['sum2'],3)
                    sum3 = round(i['sum3'],3)
                    sum4 = round(i['sum4'],3)
                    sum5 = round(i['sum5'],3)
                    sum6 = round(i['sum6'],3)
                    sum7 = round(i['sum7'],3)
                    sum8 = round(i['sum8'],3)
                    sum9 = round(i['sum9'],3)
                    sum10 = round(i['sum10'],3)
                    sum11 = round(i['sum11'],3)
                    sum12 = round(i['sum12'],3)
                    fecfin =i['fmx']
                    fecini =i['fmi']
                    diffec = fecfin - fecini
                    diffectxt = str(diffec)
                    #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                    # Crear las listas individuales para luego crear la general
                    colu1t3.append(idp)
                    colu2t3.append(prod)
                    colu3t3.append(lot)
                    colu4t3.append(sum1)
                    colu5t3.append(sum2)
                    colu6t3.append(sum3)
                    colu7t3.append(sum4)
                    colu8t3.append(sum5)
                    colu9t3.append(sum6)
                    colu10t3.append(sum7)
                    colu11t3.append(sum8)
                    colu12t3.append(sum9)
                    colu13t3.append(sum10)
                    colu14t3.append(sum11)
                    colu15t3.append(sum12)
                    colu16t3.append(fecfin)
                    colu17t3.append(fecini)
                    colu18t3.append(diffectxt)                   
            # Se crea un unica tabla
            resultado4 = zip(colu1t3,colu2t3,colu3t3,colu4t3,colu5t3,colu6t3,colu7t3,colu8t3,colu9t3,colu10t3,colu11t3,colu12t3,colu13t3,colu14t3,colu15t3,colu16t3,colu17t3,colu18t3)

        greeting = {}

        greeting['heading'] = "PROCESO DE MEZCLADO POR DIA"
        greeting['pageview'] = "FECHA : "+ fech
        greeting['fechatxt'] = fech
        greeting['no_fila1'] = len(colu1) + 1
        greeting['no_fila2'] = len(colu1t1) + 1
        greeting['no_fila3'] = len(colu1t2) + 1
        greeting['no_fila4'] = len(colu1t3) + 1                          
        greeting['sub_txt1'] = "PROCESO DE MEZCLADO"
        greeting['sub_txt2'] = "Reporte de KG # "
        greeting['table1'] = resultado1
        greeting['table2'] = resultado2
        greeting['table3'] = resultado3
        greeting['table4'] = resultado4
        greeting['table5'] = resultado_suma
        greeting['table6'] = resultado_silo                
        greeting['vfech'] = 2
        greeting['excelE'] = 1

        return render (request,'boquillas/inf_mezclado.html',greeting)
   
    # Metodo que se ejecuta antes de validar y listar los datos
    def post(self, request):
        #print('*********POST*********')
        # ********************************************    
        # Inicio
        # ******************************************** 
        now = datetime.now()  # current date and time
        # Declaracion de variables
        tz = pytz.timezone('America/Bogota') # current time zone local
        # Se lee la variable ¨fecha¨ desde el POST
        fecha = request.POST.get('fecha')
        if fecha:
            dato = datetime.strptime(fecha,"%Y%m%d")
            fech = dato.astimezone(tz).strftime("%Y-%m-%d")
        else:
            fech = now.strftime("%Y-%m-%d")
        #print("Fecha1 input: ",fecha)
        fech1 = datetime.strptime(fech,"%Y-%m-%d")
        #print('fech1: ',fech1)
        fech1z = tz.localize(fech1)
        #print('fech1z: ',fech1z)
        fech2 = fech1 + timedelta(days=1)  # day addition operation
        fech2z = tz.localize(fech2)
        #print('fech2z: ',fech2z)        
        fecht1z = tz.localize(fech1)
        fecht4z = tz.localize(fech2)
        fecht2z = fecht1z + timedelta(hours=7)  # hours addition operation
        fecht3z = fecht1z + timedelta(hours=16,minutes=30)  # hours addition operation
        # print('fecha t1: ',fecht1z)
        # print('fecha t2: ',fecht2z)
        # print('fecha t3: ',fecht3z)
        # print('fecha t4: ',fecht4z)
        # ********************************************    
        # Datos para la tabla #1
        # ********************************************
        # Se defines las variables
        colu1 = []
        colu2 = []
        colu3 = []
        colu4 = []
        colu5 = []
        colu6 = []
        colu7 = []
        colu8 = []
        colu9 = []
        colu10 = []
        colu11 = []
        colu12 = []
        colu13 = []
        colu14 = []
        colu15 = []
        colu16 = []
        colu17 = []
        colu18 = []
        colu1t1 = []
        colu2t1 = []
        colu3t1 = []
        colu4t1 = []
        colu5t1 = []
        colu6t1 = []
        colu7t1 = []
        colu8t1 = []
        colu9t1 = []
        colu10t1 = []
        colu11t1 = []
        colu12t1 = []
        colu13t1 = []
        colu14t1 = []
        colu15t1 = []
        colu16t1 = []
        colu17t1 = []
        colu18t1 = []
        colu1t2 = []
        colu2t2 = []
        colu3t2 = []
        colu4t2 = []
        colu5t2 = []
        colu6t2 = []
        colu7t2 = []
        colu8t2 = []
        colu9t2 = []
        colu10t2 = []
        colu11t2 = []
        colu12t2 = []
        colu13t2 = []
        colu14t2 = []
        colu15t2 = []
        colu16t2 = []
        colu17t2 = []
        colu18t2 = []
        colu1t3 = []
        colu2t3 = []
        colu3t3 = []
        colu4t3 = []
        colu5t3 = []
        colu6t3 = []
        colu7t3 = []
        colu8t3 = []
        colu9t3 = []
        colu10t3 = []
        colu11t3 = []
        colu12t3 = []
        colu13t3 = []
        colu14t3 = []
        colu15t3 = []
        colu16t3 = []
        colu17t3 = []
        colu18t3 = []
        silo1t1 = []
        silo2t1 = []
        silo3t1 = []
        silo4t1 = []
        silo5t1 = []
        silo6t1 = []
        silo1t2 = []
        silo2t2 = []
        silo3t2 = []
        silo4t2 = []
        silo5t2 = []
        silo6t2 = []        
        sumt1 = []
        sumt2 = []
        sumt3 = []
        sumt4 = []
        sumt5 = []
        sumt6 = []
        sumt7 = []
        sumt8 = []
        sumt9 = []
        sumt10 = []
        sumt11 = []
        sumt12 = []                        
                                                                     
        # Se guarda los productos del dia
        list_mezc = mezclado.objects.filter(datetime_plc__gte = fech1z, datetime_plc__lt = fech2z).order_by('id') 
        #for i in list_mezc:
        #   print(i.id,i.batch_id,i.lote,i.id_producto,i.name_producto,i.datetime_plc,i.bz_agre1_sp,i.bz_agre2_sp,i.bz_agre3_sp,i.bz_agre4_sp,i.bz_cem1_sp,i.bz_cem2_sp)
        # Valido si existe data
        if list_mezc:
            # Se crea una lista por producto por dia y se cuenta los SP y los PV               
            list_mezcp = list_mezc.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
            fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
            sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
            sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
            sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
            list_mezcp = list_mezcp.order_by('lote')
            for i in list_mezcp:
                #print(i)
                #print('id_prod: ',i['id_producto'])
                id = i['id_producto']
                idp = productos.objects.get(id = id)
                prod = i['name_producto']
                lot = i['lote']
                sum1 = round(i['sum1'],3)
                sum2 = round(i['sum2'],3)
                sum3 = round(i['sum3'],3)
                sum4 = round(i['sum4'],3)
                sum5 = round(i['sum5'],3)
                sum6 = round(i['sum6'],3)
                sum7 = round(i['sum7'],3)
                sum8 = round(i['sum8'],3)
                sum9 = round(i['sum9'],3)
                sum10 = round(i['sum10'],3)
                sum11 = round(i['sum11'],3)
                sum12 = round(i['sum12'],3)                
                fecfin =i['fmx']
                fecini =i['fmi']
                diffec = fecfin - fecini
                diffectxt = str(diffec)
                #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                # Crear las listas individuales para luego crear la general
                colu1.append(idp)
                colu2.append(prod)
                colu3.append(lot)
                colu4.append(sum1)
                colu5.append(sum2)
                colu6.append(sum3)
                colu7.append(sum4)
                colu8.append(sum5)
                colu9.append(sum6)
                colu10.append(sum7)
                colu11.append(sum8)
                colu12.append(sum9)
                colu13.append(sum10)
                colu14.append(sum11)
                colu15.append(sum12)
                colu16.append(fecfin)
                colu17.append(fecini)
                colu18.append(diffectxt)                   

            # Se crea una lista por year,mes y dia
            list_sumt = list_mezc.values('year_start','month_start','day_start').order_by('day_start').annotate(idf=Max('id'),
            su1=Sum('bz_agre1_sp'),su2=Sum('bz_agre1_pv'),su3=Sum('bz_agre2_sp'),su4=Sum('bz_agre2_pv'),
            su5=Sum('bz_agre3_sp'),su6=Sum('bz_agre3_pv'),su7=Sum('bz_agre4_sp'),su8=Sum('bz_agre4_pv'),
            su9=Sum('bz_cem1_sp'),su10=Sum('bz_cem1_pv'),su11=Sum('bz_cem2_sp'),su12=Sum('bz_cem2_pv'))
            # Verifico si existen datos
            if list_sumt:
                for i in list_sumt:
                    #print(i['idf'])
                    idf = i['idf']                       
                    mescf = mezclado.objects.get(id = idf)
                    #print('silos: ',mescf.silo1,mescf.silo2,mescf.silo3,mescf.silo4,mescf.silo5,mescf.silo6)
                    sumt1.append(round(i['su1'],3))
                    sumt2.append(round(i['su2'],3))
                    sumt3.append(round(i['su3'],3))
                    sumt4.append(round(i['su4'],3))
                    sumt5.append(round(i['su5'],3))
                    sumt6.append(round(i['su6'],3))
                    sumt7.append(round(i['su7'],3))
                    sumt8.append(round(i['su8'],3))
                    sumt9.append(round(i['su9'],3))
                    sumt10.append(round(i['su10'],3))
                    sumt11.append(round(i['su11'],3))
                    sumt12.append(round(i['su12'],3))
                    silo1t1.append(round(mescf.silo1,3))
                    silo1t2.append(0)
                    Nivel_metros_s1 = round(mescf.silo1,3)
                    silo2t1.append(round(mescf.silo2,3))
                    silo2t2.append(0)
                    Nivel_metros_s2 = round(mescf.silo2,3)
                    silo3t1.append(round(mescf.silo3,3))
                    silo3t2.append(0)
                    Nivel_metros_s3 = round(mescf.silo3,3)
                    silo4t1.append(round(mescf.silo4,3))
                    silo4t2.append(0)
                    Nivel_metros_s4 = round(mescf.silo4,3)
                    silo5t1.append(round(mescf.silo5,3))
                    silo5t2.append(0)
                    Nivel_metros_s5 = round(mescf.silo5,3)
                    silo6t1.append(round(mescf.silo6,3))
                    silo6t2.append(0)
                    Nivel_metros_s6 = round(mescf.silo6,3)
        # Se crea un unica tabla
        resultado1 = zip(colu1,colu2,colu3,colu4,colu5,colu6,colu7,colu8,colu9,colu10,colu11,colu12,colu13,colu14,colu15,colu16,colu17,colu18)           
        Lista_Silos_toneladas = calculo_ton_silos(Nivel_metros_s1,Nivel_metros_s2,Nivel_metros_s3,Nivel_metros_s4,Nivel_metros_s5,Nivel_metros_s6)
        #print(Lista_Silos_toneladas)
        if Lista_Silos_toneladas:
            silo1t2[0] = Lista_Silos_toneladas[0]
            silo2t2[0] = Lista_Silos_toneladas[1]
            silo3t2[0] = Lista_Silos_toneladas[2]
            silo4t2[0] = Lista_Silos_toneladas[3]
            silo5t2[0] = Lista_Silos_toneladas[4]
            silo6t2[0] = Lista_Silos_toneladas[5]
        resultado_silo = zip(silo1t1,silo1t2,silo2t1,silo2t2,silo3t1,silo3t2,silo4t1,silo4t2,silo5t1,silo5t2,silo6t1,silo6t2)  
        resultado_suma = zip(sumt1,sumt2,sumt3,sumt4,sumt5,sumt6,sumt7,sumt8,sumt9,sumt10,sumt11,sumt12)           
        
        # ****************************************************
        # Se crean las tablas de produccion por turnos del dia
        list_mezct1 = mezclado.objects.filter(datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z).order_by('id')   
        list_mezct2 = mezclado.objects.filter(datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z).order_by('id')
        list_mezct3 = mezclado.objects.filter(datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z).order_by('id')  
        # Verifico si existen datos
        if list_mezct1:
            # Se crea una lista por producto por dia y se cuenta los SP y los PV               
            list_mezct1p = list_mezct1.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
            fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
            sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
            sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
            sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
            list_mezct1p = list_mezct1p.order_by('lote')
            for i in list_mezct1p:
                #print(i)
                #print('id_prod: ',i['id_producto'])
                id = i['id_producto']
                idp = productos.objects.get(id = id)
                prod = i['name_producto']
                lot = i['lote']
                sum1 = round(i['sum1'],3)
                sum2 = round(i['sum2'],3)
                sum3 = round(i['sum3'],3)
                sum4 = round(i['sum4'],3)
                sum5 = round(i['sum5'],3)
                sum6 = round(i['sum6'],3)
                sum7 = round(i['sum7'],3)
                sum8 = round(i['sum8'],3)
                sum9 = round(i['sum9'],3)
                sum10 = round(i['sum10'],3)
                sum11 = round(i['sum11'],3)
                sum12 = round(i['sum12'],3)
                fecfin =i['fmx']
                fecini =i['fmi']
                diffec = fecfin - fecini
                diffectxt = str(diffec)
                #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                # Crear las listas individuales para luego crear la general
                colu1t1.append(idp)
                colu2t1.append(prod)
                colu3t1.append(lot)
                colu4t1.append(sum1)
                colu5t1.append(sum2)
                colu6t1.append(sum3)
                colu7t1.append(sum4)
                colu8t1.append(sum5)
                colu9t1.append(sum6)
                colu10t1.append(sum7)
                colu11t1.append(sum8)
                colu12t1.append(sum9)
                colu13t1.append(sum10)
                colu14t1.append(sum11)
                colu15t1.append(sum12)
                colu16t1.append(fecfin)
                colu17t1.append(fecini)
                colu18t1.append(diffectxt)                   
        # Se crea un unica tabla
        resultado2 = zip(colu1t1,colu2t1,colu3t1,colu4t1,colu5t1,colu6t1,colu7t1,colu8t1,colu9t1,colu10t1,colu11t1,colu12t1,colu13t1,colu14t1,colu15t1,colu16t1,colu17t1,colu18t1)

        # Verifico si existen datos
        if list_mezct2:
            # Se crea una lista por producto por dia y se cuenta los SP y los PV               
            list_mezct2p = list_mezct2.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
            fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
            sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
            sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
            sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
            list_mezct2p = list_mezct2p.order_by('lote')
            for i in list_mezct2p:
                #print(i)
                #print('id_prod: ',i['id_producto'])
                id = i['id_producto']
                idp = productos.objects.get(id = id)
                prod = i['name_producto']
                lot = i['lote']
                sum1 = round(i['sum1'],3)
                sum2 = round(i['sum2'],3)
                sum3 = round(i['sum3'],3)
                sum4 = round(i['sum4'],3)
                sum5 = round(i['sum5'],3)
                sum6 = round(i['sum6'],3)
                sum7 = round(i['sum7'],3)
                sum8 = round(i['sum8'],3)
                sum9 = round(i['sum9'],3)
                sum10 = round(i['sum10'],3)
                sum11 = round(i['sum11'],3)
                sum12 = round(i['sum12'],3)
                fecfin =i['fmx']
                fecini =i['fmi']
                diffec = fecfin - fecini
                diffectxt = str(diffec)
                #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                # Crear las listas individuales para luego crear la general
                colu1t2.append(idp)
                colu2t2.append(prod)
                colu3t2.append(lot)
                colu4t2.append(sum1)
                colu5t2.append(sum2)
                colu6t2.append(sum3)
                colu7t2.append(sum4)
                colu8t2.append(sum5)
                colu9t2.append(sum6)
                colu10t2.append(sum7)
                colu11t2.append(sum8)
                colu12t2.append(sum9)
                colu13t2.append(sum10)
                colu14t2.append(sum11)
                colu15t2.append(sum12)
                colu16t2.append(fecfin)
                colu17t2.append(fecini)
                colu18t2.append(diffectxt)                   
        # Se crea un unica tabla
        resultado3 = zip(colu1t2,colu2t2,colu3t2,colu4t2,colu5t2,colu6t2,colu7t2,colu8t2,colu9t2,colu10t2,colu11t2,colu12t2,colu13t2,colu14t2,colu15t2,colu16t2,colu17t2,colu18t2)

        # Verifico si existen datos
        if list_mezct3:
            # Se crea una lista por producto por dia y se cuenta los SP y los PV               
            list_mezct3p = list_mezct3.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
            fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
            sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
            sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
            sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
            list_mezct3p = list_mezct3p.order_by('lote')
            for i in list_mezct3p:
                #print(i)
                #print('id_prod: ',i['id_producto'])
                id = i['id_producto']
                idp = productos.objects.get(id = id)
                prod = i['name_producto']
                lot = i['lote']
                sum1 = round(i['sum1'],3)
                sum2 = round(i['sum2'],3)
                sum3 = round(i['sum3'],3)
                sum4 = round(i['sum4'],3)
                sum5 = round(i['sum5'],3)
                sum6 = round(i['sum6'],3)
                sum7 = round(i['sum7'],3)
                sum8 = round(i['sum8'],3)
                sum9 = round(i['sum9'],3)
                sum10 = round(i['sum10'],3)
                sum11 = round(i['sum11'],3)
                sum12 = round(i['sum12'],3)
                fecfin =i['fmx']
                fecini =i['fmi']
                diffec = fecfin - fecini
                diffectxt = str(diffec)
                #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                # Crear las listas individuales para luego crear la general
                colu1t3.append(idp)
                colu2t3.append(prod)
                colu3t3.append(lot)
                colu4t3.append(sum1)
                colu5t3.append(sum2)
                colu6t3.append(sum3)
                colu7t3.append(sum4)
                colu8t3.append(sum5)
                colu9t3.append(sum6)
                colu10t3.append(sum7)
                colu11t3.append(sum8)
                colu12t3.append(sum9)
                colu13t3.append(sum10)
                colu14t3.append(sum11)
                colu15t3.append(sum12)
                colu16t3.append(fecfin)
                colu17t3.append(fecini)
                colu18t3.append(diffectxt)                   
        # Se crea un unica tabla
        resultado4 = zip(colu1t3,colu2t3,colu3t3,colu4t3,colu5t3,colu6t3,colu7t3,colu8t3,colu9t3,colu10t3,colu11t3,colu12t3,colu13t3,colu14t3,colu15t3,colu16t3,colu17t3,colu18t3)
         
        greeting = {}

        greeting['heading'] = "PROCESO DE MEZCLADO POR DIA"
        greeting['pageview'] = "FECHA : "+ fech
        greeting['fechatxt'] = fech
        greeting['no_fila1'] = len(colu1) + 1
        greeting['no_fila2'] = len(colu1t1) + 1
        greeting['no_fila3'] = len(colu1t2) + 1
        greeting['no_fila4'] = len(colu1t3) + 1       
        greeting['sub_txt1'] = "PROCESO DE MEZCLADO"
        greeting['sub_txt2'] = "Reporte de KG # "
        greeting['table1'] = resultado1
        greeting['table2'] = resultado2
        greeting['table3'] = resultado3
        greeting['table4'] = resultado4
        greeting['table5'] = resultado_suma
        greeting['table6'] = resultado_silo         
        greeting['vfech'] = 2
        greeting['excelE'] = 1

        return render (request,'boquillas/inf_mezclado.html',greeting) 

# ********************************************************************************************
# ********************************************************************************************
# Generador de documento en Excel y envio por correo =========================================
class CrearArchivoExcel(View):
    
    def get(self, request, fecha = None, excelE = None):
        #print('*********GET*********')
        # ********************************************    
        # Se crea la funcion de Excel
        # ********************************************
        # Se crea la funcion de Excel
        def excel(fecha):
            
            from openpyxl import load_workbook

            FILE_PATH ='C:/inetpub/wwwroot/argos_ml_project/static/excel/Informe_Q&PTECH_Panama_Pantilla.xlsx'
            sheet3 = 'InformeQ&PTECHPanama'
            
            workbook= load_workbook(FILE_PATH,read_only=False)
            ws3=workbook[sheet3]

            # Proceso de ensacado
            # Colocar la fecha en la Fila 1 Columna A
            ws3.cell(row = 2,column = 1, value = fech)
            # Carga de la Columna B
            rowIndex = 4
            for filas in colu1e:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 2, value = int(str(filas)))
                rowIndex +=1
            # Carga de la Columna C
            rowIndex = 4
            for filas in colu2e:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 3, value = filas)
                rowIndex +=1
            # Carga de la Columna D
            rowIndex = 4
            for filas in colu3e:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 4, value = filas)
                rowIndex +=1
            # Carga de la Columna E
            rowIndex = 4
            for filas in colu4e:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 5, value = filas)
                rowIndex +=1

            # Proceso de mezclado
            # Carga de la Columna B
            rowIndex = 18
            for filas in colu1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 2, value = int(str(filas)))
                rowIndex +=1
            # Carga de la Columna C
            rowIndex = 18
            for filas in colu2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 3, value = filas)
                rowIndex +=1
            # Carga de la Columna D
            rowIndex = 18
            for filas in colu3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 4, value = filas)
                rowIndex +=1
            # Carga de la Columna E
            rowIndex = 18
            for filas in colu4:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 5, value = filas)
                rowIndex +=1
            # Carga de la Columna F
            rowIndex = 18
            for filas in colu5:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 6, value = filas)
                rowIndex +=1
            # Carga de la Columna G
            rowIndex = 18
            for filas in colu6:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 7, value = filas)
                rowIndex +=1
            # Carga de la Columna H
            rowIndex = 18
            for filas in colu7:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 8, value = filas)
                rowIndex +=1
            # Carga de la Columna I
            rowIndex = 18
            for filas in colu8:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 9, value = filas)
                rowIndex +=1
            # Carga de la Columna J
            rowIndex = 18
            for filas in colu9:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 10, value = filas)
                rowIndex +=1
            # Carga de la Columna K
            rowIndex = 18
            for filas in colu10:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 11, value = filas)
                rowIndex +=1
            # Carga de la Columna L
            rowIndex = 18
            for filas in colu11:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 12, value = filas)
                rowIndex +=1
            # Carga de la Columna M
            rowIndex = 18
            for filas in colu12:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 13, value = filas)
                rowIndex +=1
            # Carga de la Columna N
            rowIndex = 18
            for filas in colu13:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 14, value = filas)
                rowIndex +=1
            # Carga de la Columna O
            rowIndex = 18
            for filas in colu14:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 15, value = filas)
                rowIndex +=1
            # Carga de la Columna P
            rowIndex = 18
            for filas in colu15:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 16, value = filas)
                rowIndex +=1
            # Carga de la Columna Q
            rowIndex = 18
            for filas in colu17:
                # Cargo el dato en la Celda
                dataf = filas - timedelta(hours=5)
                ws3.cell(row = rowIndex,column = 17, value = dataf.strftime("%Y-%m-%d %H:%M:%S"))
                rowIndex +=1
            # Carga de la Columna R
            rowIndex = 18
            for filas in colu16:
                # Cargo el dato en la Celda
                dataf = filas - timedelta(hours=5)
                ws3.cell(row = rowIndex,column = 18, value = dataf.strftime("%Y-%m-%d %H:%M:%S"))
                rowIndex +=1
            # Carga de la Columna S
            rowIndex = 18
            for filas in colu18:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 19, value = filas)
                rowIndex +=1

            # Carga de la Columna E
            rowIndex = 27
            for filas in sumt1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 5, value = filas)
                rowIndex +=1
            # Carga de la Columna F
            rowIndex = 27
            for filas in sumt2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 6, value = filas)
                rowIndex +=1
            # Carga de la Columna G
            rowIndex = 27
            for filas in sumt3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 7, value = filas)
                rowIndex +=1
            # Carga de la Columna H
            rowIndex = 27
            for filas in sumt4:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 8, value = filas)
                rowIndex +=1
            # Carga de la Columna I
            rowIndex = 27
            for filas in sumt5:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 9, value = filas)
                rowIndex +=1
            # Carga de la Columna J
            rowIndex = 27
            for filas in sumt6:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 10, value = filas)
                rowIndex +=1
            # Carga de la Columna K
            rowIndex = 27
            for filas in sumt7:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 11, value = filas)
                rowIndex +=1
            # Carga de la Columna L
            rowIndex = 27
            for filas in sumt8:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 12, value = filas)
                rowIndex +=1
            # Carga de la Columna M
            rowIndex = 27
            for filas in sumt9:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 13, value = filas)
                rowIndex +=1
            # Carga de la Columna N
            rowIndex = 27
            for filas in sumt10:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 14, value = filas)
                rowIndex +=1
            # Carga de la Columna O
            rowIndex = 27
            for filas in sumt11:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 15, value = filas)
                rowIndex +=1
            # Carga de la Columna P
            rowIndex = 27
            for filas in sumt12:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 16, value = filas)
                rowIndex +=1

            # Carga de la Columna B
            rowIndex = 33
            for filas in colu1t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 2, value = int(str(filas)))
                rowIndex +=1
            # Carga de la Columna C
            rowIndex = 33
            for filas in colu2t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 3, value = filas)
                rowIndex +=1
            # Carga de la Columna D
            rowIndex = 33
            for filas in colu3t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 4, value = filas)
                rowIndex +=1
            # Carga de la Columna E
            rowIndex = 33
            for filas in colu4t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 5, value = filas)
                rowIndex +=1
            # Carga de la Columna F
            rowIndex = 33
            for filas in colu5t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 6, value = filas)
                rowIndex +=1
            # Carga de la Columna G
            rowIndex = 33
            for filas in colu6t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 7, value = filas)
                rowIndex +=1
            # Carga de la Columna H
            rowIndex = 33
            for filas in colu7t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 8, value = filas)
                rowIndex +=1
            # Carga de la Columna I
            rowIndex = 33
            for filas in colu8t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 9, value = filas)
                rowIndex +=1
            # Carga de la Columna J
            rowIndex = 33
            for filas in colu9t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 10, value = filas)
                rowIndex +=1
            # Carga de la Columna K
            rowIndex = 33
            for filas in colu10t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 11, value = filas)
                rowIndex +=1
            # Carga de la Columna L
            rowIndex = 33
            for filas in colu11t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 12, value = filas)
                rowIndex +=1
            # Carga de la Columna M
            rowIndex = 33
            for filas in colu12t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 13, value = filas)
                rowIndex +=1
            # Carga de la Columna N
            rowIndex = 33
            for filas in colu13t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 14, value = filas)
                rowIndex +=1
            # Carga de la Columna O
            rowIndex = 33
            for filas in colu14t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 15, value = filas)
                rowIndex +=1
            # Carga de la Columna P
            rowIndex = 33
            for filas in colu15t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 16, value = filas)
                rowIndex +=1
            # Carga de la Columna Q
            rowIndex = 33
            for filas in colu17t1:
                # Cargo el dato en la Celda
                dataf = filas - timedelta(hours=5)
                ws3.cell(row = rowIndex,column = 17, value = dataf.strftime("%Y-%m-%d %H:%M:%S"))
                rowIndex +=1
            # Carga de la Columna R
            rowIndex = 33
            for filas in colu16t1:
                # Cargo el dato en la Celda
                dataf = filas - timedelta(hours=5)
                ws3.cell(row = rowIndex,column = 18, value = dataf.strftime("%Y-%m-%d %H:%M:%S"))
                rowIndex +=1
            # Carga de la Columna S
            rowIndex = 33
            for filas in colu18t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 19, value = filas)
                rowIndex +=1

            # Carga de la Columna B
            rowIndex = 37
            for filas in colu1t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 2, value = int(str(filas)))
                rowIndex +=1
            # Carga de la Columna C
            rowIndex = 37
            for filas in colu2t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 3, value = filas)
                rowIndex +=1
            # Carga de la Columna D
            rowIndex = 37
            for filas in colu3t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 4, value = filas)
                rowIndex +=1
            # Carga de la Columna E
            rowIndex = 37
            for filas in colu4t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 5, value = filas)
                rowIndex +=1
            # Carga de la Columna F
            rowIndex = 37
            for filas in colu5t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 6, value = filas)
                rowIndex +=1
            # Carga de la Columna G
            rowIndex = 37
            for filas in colu6t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 7, value = filas)
                rowIndex +=1
            # Carga de la Columna H
            rowIndex = 37
            for filas in colu7t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 8, value = filas)
                rowIndex +=1
            # Carga de la Columna I
            rowIndex = 37
            for filas in colu8t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 9, value = filas)
                rowIndex +=1
            # Carga de la Columna J
            rowIndex = 37
            for filas in colu9t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 10, value = filas)
                rowIndex +=1
            # Carga de la Columna K
            rowIndex = 37
            for filas in colu10t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 11, value = filas)
                rowIndex +=1
            # Carga de la Columna L
            rowIndex = 37
            for filas in colu11t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 12, value = filas)
                rowIndex +=1
            # Carga de la Columna M
            rowIndex = 37
            for filas in colu12t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 13, value = filas)
                rowIndex +=1
            # Carga de la Columna N
            rowIndex = 37
            for filas in colu13t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 14, value = filas)
                rowIndex +=1
            # Carga de la Columna O
            rowIndex = 37
            for filas in colu14t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 15, value = filas)
                rowIndex +=1
            # Carga de la Columna P
            rowIndex = 37
            for filas in colu15t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 16, value = filas)
                rowIndex +=1
            # Carga de la Columna Q
            rowIndex = 37
            for filas in colu17t2:
                # Cargo el dato en la Celda
                dataf = filas - timedelta(hours=5)
                ws3.cell(row = rowIndex,column = 17, value = dataf.strftime("%Y-%m-%d %H:%M:%S"))
                rowIndex +=1
            # Carga de la Columna R
            rowIndex = 37
            for filas in colu16t2:
                # Cargo el dato en la Celda
                dataf = filas - timedelta(hours=5)
                ws3.cell(row = rowIndex,column = 18, value = dataf.strftime("%Y-%m-%d %H:%M:%S"))
                rowIndex +=1
            # Carga de la Columna S
            rowIndex = 37
            for filas in colu18t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 19, value = filas)
                rowIndex +=1

            # Carga de la Columna B
            rowIndex = 41
            for filas in colu1t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 2, value = int(str(filas)))
                rowIndex +=1
            # Carga de la Columna C
            rowIndex = 41
            for filas in colu2t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 3, value = filas)
                rowIndex +=1
            # Carga de la Columna D
            rowIndex = 41
            for filas in colu3t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 4, value = filas)
                rowIndex +=1
            # Carga de la Columna E
            rowIndex = 41
            for filas in colu4t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 5, value = filas)
                rowIndex +=1
            # Carga de la Columna F
            rowIndex = 41
            for filas in colu5t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 6, value = filas)
                rowIndex +=1
            # Carga de la Columna G
            rowIndex = 41
            for filas in colu6t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 7, value = filas)
                rowIndex +=1
            # Carga de la Columna H
            rowIndex = 41
            for filas in colu7t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 8, value = filas)
                rowIndex +=1
            # Carga de la Columna I
            rowIndex = 41
            for filas in colu8t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 9, value = filas)
                rowIndex +=1
            # Carga de la Columna J
            rowIndex = 41
            for filas in colu9t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 10, value = filas)
                rowIndex +=1
            # Carga de la Columna K
            rowIndex = 41
            for filas in colu10t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 11, value = filas)
                rowIndex +=1
            # Carga de la Columna L
            rowIndex = 41
            for filas in colu11t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 12, value = filas)
                rowIndex +=1
            # Carga de la Columna M
            rowIndex = 41
            for filas in colu12t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 13, value = filas)
                rowIndex +=1
            # Carga de la Columna N
            rowIndex = 41
            for filas in colu13t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 14, value = filas)
                rowIndex +=1
            # Carga de la Columna O
            rowIndex = 41
            for filas in colu14t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 15, value = filas)
                rowIndex +=1
            # Carga de la Columna P
            rowIndex = 41
            for filas in colu15t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 16, value = filas)
                rowIndex +=1
            # Carga de la Columna Q
            rowIndex = 41
            for filas in colu17t3:
                # Cargo el dato en la Celda
                dataf = filas - timedelta(hours=5)
                ws3.cell(row = rowIndex,column = 17, value = dataf.strftime("%Y-%m-%d %H:%M:%S"))
                rowIndex +=1
            # Carga de la Columna R
            rowIndex = 41
            for filas in colu16t3:
                # Cargo el dato en la Celda
                dataf = filas - timedelta(hours=5)
                ws3.cell(row = rowIndex,column = 18, value = dataf.strftime("%Y-%m-%d %H:%M:%S"))
                rowIndex +=1
            # Carga de la Columna S
            rowIndex = 41
            for filas in colu18t3:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 19, value = filas)
                rowIndex +=1

            # Proceso por silo
            # Carga de la Columna A
            rowIndex = 50
            for filas in silo1t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 1, value = filas)
                rowIndex +=1
            # Carga de la Columna B
            rowIndex = 50
            for filas in silo1t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 2, value = filas)
                rowIndex +=1
            # Carga de la Columna C
            rowIndex = 50
            for filas in silo2t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 3, value = filas)
                rowIndex +=1
            # Carga de la Columna D
            rowIndex = 50
            for filas in silo2t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 4, value = filas)
                rowIndex +=1
            # Carga de la Columna E
            rowIndex = 50
            for filas in silo3t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 5, value = filas)
                rowIndex +=1
            # Carga de la Columna F
            rowIndex = 50
            for filas in silo3t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 6, value = filas)
                rowIndex +=1
            # Carga de la Columna G
            rowIndex = 50
            for filas in silo4t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 7, value = filas)
                rowIndex +=1
            # Carga de la Columna H
            rowIndex = 50
            for filas in silo4t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 8, value = filas)
                rowIndex +=1
            # Carga de la Columna I
            rowIndex = 50
            for filas in silo5t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 9, value = filas)
                rowIndex +=1
            # Carga de la Columna J
            rowIndex = 50
            for filas in silo5t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 10, value = filas)
                rowIndex +=1
            # Carga de la Columna K
            rowIndex = 50
            for filas in silo6t1:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 11, value = filas)
                rowIndex +=1
            # Carga de la Columna L
            rowIndex = 50
            for filas in silo6t2:
                # Cargo el dato en la Celda
                ws3.cell(row = rowIndex,column = 12, value = filas)
                rowIndex +=1

            fileName = f"C:/inetpub/wwwroot/argos_ml_project/static/excel/InformeEnsacadoMezclado.xlsx"
            
            workbook.save(fileName)  
            return fileName

        # ********************************************    
        # Inicio
        # ******************************************** 
        now = datetime.now()  # current date and time
        # Declaracion de variables
        tz = pytz.timezone('America/Bogota') # current time zone local
        year = date.today().year # current year
        month = date.today().month # current month
        day = date.today().day # current day
        fecha0 = list()
        fecha0 = ['01','01','2024']
        fecha0[0] = day
        fecha0[1] = month
        fecha0[2] = year
        file_name = str(f"InformeEnsacadoMezclado_{fecha0[0]}-{fecha0[1]}-{fecha0[2]}.xlsx")
        file_path = str(f"C:/inetpub/wwwroot/argos_ml_project/static/excel/InformeEnsacadoMezclado_{fecha0[0]}-{fecha0[1]}-{fecha0[2]}.xlsx")
        # Se lee la variable ¨fecha¨ desde el GET
        #print('fecha: ',fecha)
        if fecha:
            dato = datetime.strptime(fecha,"%Y-%m-%d")
            fech = dato.astimezone(tz).strftime("%Y-%m-%d")
        else:
            fech = now.strftime("%Y-%m-%d")        
        fech1 = datetime.strptime(fech,"%Y-%m-%d")
        #print('fech1: ',fech1)
        fech1z = tz.localize(fech1)
        #print('fech1z: ',fech1z)
        fech2 = fech1 + timedelta(days=1)  # day addition operation
        fech2z = tz.localize(fech2)
        #print('fech2z: ',fech2z)        
        fecht1z = tz.localize(fech1)
        fecht4z = tz.localize(fech2)
        fecht2z = fecht1z + timedelta(hours=7)  # hours addition operation
        fecht3z = fecht1z + timedelta(hours=16,minutes=30)  # hours addition operation
        #print('fecha t1: ',fecht1z)
        #print('fecha t2: ',fecht2z)
        #print('fecha t3: ',fecht3z)
        #print('fecha t4: ',fecht4z)
        # ********************************************    
        # Datos para la tabla #1
        # ********************************************
        # Se defines las variables
        colu1e = []
        colu2e = []
        colu3e = []
        colu4e = []
        colu1 = []
        colu2 = []
        colu3 = []
        colu4 = []
        colu5 = []
        colu6 = []
        colu7 = []
        colu8 = []
        colu9 = []
        colu10 = []
        colu11 = []
        colu12 = []
        colu13 = []
        colu14 = []
        colu15 = []
        colu16 = []
        colu17 = []
        colu18 = []
        silo1t1 = []
        silo2t1 = []
        silo3t1 = []
        silo4t1 = []
        silo5t1 = []
        silo6t1 = []
        silo1t2 = []
        silo2t2 = []
        silo3t2 = []
        silo4t2 = []
        silo5t2 = []
        silo6t2 = []
        sumt1 = []
        sumt2 = []
        sumt3 = []
        sumt4 = []
        sumt5 = []
        sumt6 = []
        sumt7 = []
        sumt8 = []
        sumt9 = []
        sumt10 = []
        sumt11 = []
        sumt12 = []
        colu1t1 = []
        colu2t1 = []
        colu3t1 = []
        colu4t1 = []
        colu5t1 = []
        colu6t1 = []
        colu7t1 = []
        colu8t1 = []
        colu9t1 = []
        colu10t1 = []
        colu11t1 = []
        colu12t1 = []
        colu13t1 = []
        colu14t1 = []
        colu15t1 = []
        colu16t1 = []
        colu17t1 = []
        colu18t1 = []
        colu1t2 = []
        colu2t2 = []
        colu3t2 = []
        colu4t2 = []
        colu5t2 = []
        colu6t2 = []
        colu7t2 = []
        colu8t2 = []
        colu9t2 = []
        colu10t2 = []
        colu11t2 = []
        colu12t2 = []
        colu13t2 = []
        colu14t2 = []
        colu15t2 = []
        colu16t2 = []
        colu17t2 = []
        colu18t2 = []
        colu1t3 = []
        colu2t3 = []
        colu3t3 = []
        colu4t3 = []
        colu5t3 = []
        colu6t3 = []
        colu7t3 = []
        colu8t3 = []
        colu9t3 = []
        colu10t3 = []
        colu11t3 = []
        colu12t3 = []
        colu13t3 = []
        colu14t3 = []
        colu15t3 = []
        colu16t3 = []
        colu17t3 = []
        colu18t3 = []
        # ********************************************
        # Se guarda los productos del dia (Ensacado)
        # ********************************************
        list_ens = boquillas.objects.filter(datetime_plc__gte = fech1z, datetime_plc__lt = fech2z).order_by('id')

        if list_ens:
            # Se crea una lista por producto por dia y se cuenta               
            list_ensp = list_ens.values('id_producto').order_by('id_producto').annotate(count=Count('id_producto'))
            for i in list_ensp:
                #print('id_prod: ',i['id_producto'])
                idp = i['id_producto']
                prod = productos.objects.get(id = idp )
                peso = float(str(prod.peso))
                #print(peso)
                sum_saco = 0
                # Busco el ultimo numero de sacos y el primer numero de sacos por boqillas
                lsb1t1l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                lsb1t1f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                lsb1t2l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                lsb1t2f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                lsb1t3l = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                lsb1t3f = boquillas.objects.filter(n_boquilla = 1, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()

                lsb2t1l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                lsb2t1f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                lsb2t2l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                lsb2t2f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                lsb2t3l = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                lsb2t3f = boquillas.objects.filter(n_boquilla = 2, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()

                lsb3t1l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').last()
                lsb3t1f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z, id_producto = prod).order_by('id').first()
                lsb3t2l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').last()
                lsb3t2f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z, id_producto = prod).order_by('id').first()
                lsb3t3l = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').last()
                lsb3t3f = boquillas.objects.filter(n_boquilla = 3, datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z, id_producto = prod).order_by('id').first()                                       
                # Verifico si existen datos
                if lsb1t1l and lsb1t1f:
                    sum_saco = sum_saco + (lsb1t1l.idbq - lsb1t1f.idbq + 1)
                    #print('suma1: ',sum_saco)
                if lsb1t2l and lsb1t2f:
                    sum_saco = sum_saco + (lsb1t2l.idbq - lsb1t2f.idbq + 1)
                    #print('suma2: ',sum_saco)
                if lsb1t3l and lsb1t3f:
                    sum_saco = sum_saco + (lsb1t3l.idbq - lsb1t3f.idbq + 1)
                    #print('suma3: ',sum_saco)

                if lsb2t1l and lsb2t1f:
                    sum_saco = sum_saco + (lsb2t1l.idbq - lsb2t1f.idbq + 1)
                    #print('suma4: ',sum_saco)
                if lsb2t2l and lsb2t2f:
                    sum_saco = sum_saco + (lsb2t2l.idbq - lsb2t2f.idbq + 1)
                    #print('suma5: ',sum_saco)
                if lsb2t3l and lsb2t3f:
                    sum_saco = sum_saco + (lsb2t3l.idbq - lsb2t3f.idbq + 1)
                    #print('suma6: ',sum_saco)

                if lsb3t1l and lsb3t1f:
                    sum_saco = sum_saco + (lsb3t1l.idbq - lsb3t1f.idbq + 1)
                    #print('suma7: ',sum_saco)
                if lsb3t2l and lsb3t2f:
                    sum_saco = sum_saco + (lsb3t2l.idbq - lsb3t2f.idbq + 1)
                    #print('suma8: ',sum_saco)
                if lsb3t3l and lsb3t3f:
                    sum_saco = sum_saco + (lsb3t3l.idbq - lsb3t3f.idbq + 1)
                    #print('suma9: ',sum_saco)                
                ton_saco = round((sum_saco * peso)/1000,1)
                if sum_saco > 1:
                    colu1e.append(prod.id_Producto)
                    colu2e.append(prod.name)
                    colu3e.append(sum_saco)
                    colu4e.append(ton_saco)

        #for i in zip(colu1e,colu2e,colu3e,colu4e):
        #    print(i)
        # Se crea un unica tabla
        resultado1e = zip(colu1e,colu2e,colu3e,colu4e)

        # ********************************************
        # Se guarda los productos del dia (Mezclado)
        # ********************************************
        list_mezc = mezclado.objects.filter(datetime_plc__gte = fech1z, datetime_plc__lt = fech2z).order_by('id') 
        #for i in list_mezc:
        #   print(i.id,i.batch_id,i.lote,i.id_producto,i.name_producto,i.datetime_plc,i.bz_agre1_sp,i.bz_agre2_sp,i.bz_agre3_sp,i.bz_agre4_sp,i.bz_cem1_sp,i.bz_cem2_sp)
        # Valido si existe data
        if list_mezc:
            # Se crea una lista por producto por dia y se cuenta los SP y los PV               
            list_mezcp = list_mezc.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
            fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
            sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
            sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
            sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
            list_mezcp = list_mezcp.order_by('lote')
            for i in list_mezcp:
                #print(i)
                #print('id_prod: ',i['id_producto'])
                id = i['id_producto']
                idp = productos.objects.get(id = id)
                prod = i['name_producto']
                lot = i['lote']
                sum1 = round(i['sum1'],3)
                sum2 = round(i['sum2'],3)
                sum3 = round(i['sum3'],3)
                sum4 = round(i['sum4'],3)
                sum5 = round(i['sum5'],3)
                sum6 = round(i['sum6'],3)
                sum7 = round(i['sum7'],3)
                sum8 = round(i['sum8'],3)
                sum9 = round(i['sum9'],3)
                sum10 = round(i['sum10'],3)
                sum11 = round(i['sum11'],3)
                sum12 = round(i['sum12'],3)                
                fecfin =i['fmx']
                fecini =i['fmi']
                diffec = fecfin - fecini
                diffectxt = str(diffec)
                #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                # Crear las listas individuales para luego crear la general
                colu1.append(idp)
                colu2.append(prod)
                colu3.append(lot)
                colu4.append(sum1)
                colu5.append(sum2)
                colu6.append(sum3)
                colu7.append(sum4)
                colu8.append(sum5)
                colu9.append(sum6)
                colu10.append(sum7)
                colu11.append(sum8)
                colu12.append(sum9)
                colu13.append(sum10)
                colu14.append(sum11)
                colu15.append(sum12)
                colu16.append(fecfin)
                colu17.append(fecini)
                colu18.append(diffectxt)                   

            # Se crea una lista por year,mes y dia
            list_sumt = list_mezc.values('year_start','month_start','day_start').order_by('day_start').annotate(idf=Max('id'),
            su1=Sum('bz_agre1_sp'),su2=Sum('bz_agre1_pv'),su3=Sum('bz_agre2_sp'),su4=Sum('bz_agre2_pv'),
            su5=Sum('bz_agre3_sp'),su6=Sum('bz_agre3_pv'),su7=Sum('bz_agre4_sp'),su8=Sum('bz_agre4_pv'),
            su9=Sum('bz_cem1_sp'),su10=Sum('bz_cem1_pv'),su11=Sum('bz_cem2_sp'),su12=Sum('bz_cem2_pv'))
            # Verifico si existen datos
            if list_sumt:
                for i in list_sumt:
                    #print(i['idf'])
                    idf = i['idf']                       
                    mescf = mezclado.objects.get(id = idf)
                    #print('silos: ',mescf.silo1,mescf.silo2,mescf.silo3,mescf.silo4,mescf.silo5,mescf.silo6)
                    sumt1.append(round(i['su1'],3))
                    sumt2.append(round(i['su2'],3))
                    sumt3.append(round(i['su3'],3))
                    sumt4.append(round(i['su4'],3))
                    sumt5.append(round(i['su5'],3))
                    sumt6.append(round(i['su6'],3))
                    sumt7.append(round(i['su7'],3))
                    sumt8.append(round(i['su8'],3))
                    sumt9.append(round(i['su9'],3))
                    sumt10.append(round(i['su10'],3))
                    sumt11.append(round(i['su11'],3))
                    sumt12.append(round(i['su12'],3))
                    silo1t1.append(round(mescf.silo1,3))
                    silo1t2.append(0)
                    Nivel_metros_s1 = round(mescf.silo1,3)
                    silo2t1.append(round(mescf.silo2,3))
                    silo2t2.append(0)
                    Nivel_metros_s2 = round(mescf.silo2,3)
                    silo3t1.append(round(mescf.silo3,3))
                    silo3t2.append(0)
                    Nivel_metros_s3 = round(mescf.silo3,3)
                    silo4t1.append(round(mescf.silo4,3))
                    silo4t2.append(0)
                    Nivel_metros_s4 = round(mescf.silo4,3)
                    silo5t1.append(round(mescf.silo5,3))
                    silo5t2.append(0)
                    Nivel_metros_s5 = round(mescf.silo5,3)
                    silo6t1.append(round(mescf.silo6,3))
                    silo6t2.append(0)
                    Nivel_metros_s6 = round(mescf.silo6,3)
        # Se crea un unica tabla
        resultado1 = zip(colu1,colu2,colu3,colu4,colu5,colu6,colu7,colu8,colu9,colu10,colu11,colu12,colu13,colu14,colu15,colu16,colu17,colu18)           
        Lista_Silos_toneladas = calculo_ton_silos(Nivel_metros_s1,Nivel_metros_s2,Nivel_metros_s3,Nivel_metros_s4,Nivel_metros_s5,Nivel_metros_s6)
        #print(Lista_Silos_toneladas)
        if Lista_Silos_toneladas:
            silo1t2[0] = Lista_Silos_toneladas[0]
            silo2t2[0] = Lista_Silos_toneladas[1]
            silo3t2[0] = Lista_Silos_toneladas[2]
            silo4t2[0] = Lista_Silos_toneladas[3]
            silo5t2[0] = Lista_Silos_toneladas[4]
            silo6t2[0] = Lista_Silos_toneladas[5]
        resultado_silo = zip(silo1t1,silo1t2,silo2t1,silo2t2,silo3t1,silo3t2,silo4t1,silo4t2,silo5t1,silo5t2,silo6t1,silo6t2)  
        resultado_suma = zip(sumt1,sumt2,sumt3,sumt4,sumt5,sumt6,sumt7,sumt8,sumt9,sumt10,sumt11,sumt12)            
        # ****************************************************
        # Se crean las tablas de produccion por turnos del dia (Mezcla)
        # ****************************************************
        list_mezct1 = mezclado.objects.filter(datetime_plc__gte = fecht1z, datetime_plc__lt = fecht2z).order_by('id')   
        list_mezct2 = mezclado.objects.filter(datetime_plc__gte = fecht2z, datetime_plc__lt = fecht3z).order_by('id')
        list_mezct3 = mezclado.objects.filter(datetime_plc__gte = fecht3z, datetime_plc__lt = fecht4z).order_by('id')  
        # Verifico si existen datos
        if list_mezct1:
            # Se crea una lista por producto por dia y se cuenta los SP y los PV               
            list_mezct1p = list_mezct1.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
            fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
            sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
            sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
            sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
            list_mezct1p = list_mezct1p.order_by('lote')
            for i in list_mezct1p:
                #print(i)
                #print('id_prod: ',i['id_producto'])
                id = i['id_producto']
                idp = productos.objects.get(id = id)
                prod = i['name_producto']
                lot = i['lote']
                sum1 = round(i['sum1'],3)
                sum2 = round(i['sum2'],3)
                sum3 = round(i['sum3'],3)
                sum4 = round(i['sum4'],3)
                sum5 = round(i['sum5'],3)
                sum6 = round(i['sum6'],3)
                sum7 = round(i['sum7'],3)
                sum8 = round(i['sum8'],3)
                sum9 = round(i['sum9'],3)
                sum10 = round(i['sum10'],3)
                sum11 = round(i['sum11'],3)
                sum12 = round(i['sum12'],3)
                fecfin =i['fmx']
                fecini =i['fmi']
                diffec = fecfin - fecini
                diffectxt = str(diffec)
                #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                # Crear las listas individuales para luego crear la general
                colu1t1.append(idp)
                colu2t1.append(prod)
                colu3t1.append(lot)
                colu4t1.append(sum1)
                colu5t1.append(sum2)
                colu6t1.append(sum3)
                colu7t1.append(sum4)
                colu8t1.append(sum5)
                colu9t1.append(sum6)
                colu10t1.append(sum7)
                colu11t1.append(sum8)
                colu12t1.append(sum9)
                colu13t1.append(sum10)
                colu14t1.append(sum11)
                colu15t1.append(sum12)
                colu16t1.append(fecfin)
                colu17t1.append(fecini)
                colu18t1.append(diffectxt)                   
        # Se crea un unica tabla
        resultado2 = zip(colu1t1,colu2t1,colu3t1,colu4t1,colu5t1,colu6t1,colu7t1,colu8t1,colu9t1,colu10t1,colu11t1,colu12t1,colu13t1,colu14t1,colu15t1,colu16t1,colu17t1,colu18t1)

        # Verifico si existen datos
        if list_mezct2:
            # Se crea una lista por producto por dia y se cuenta los SP y los PV               
            list_mezct2p = list_mezct2.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
            fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
            sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
            sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
            sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
            list_mezct2p = list_mezct2p.order_by('lote')
            for i in list_mezct2p:
                #print(i)
                #print('id_prod: ',i['id_producto'])
                id = i['id_producto']
                idp = productos.objects.get(id = id)
                prod = i['name_producto']
                lot = i['lote']
                sum1 = round(i['sum1'],3)
                sum2 = round(i['sum2'],3)
                sum3 = round(i['sum3'],3)
                sum4 = round(i['sum4'],3)
                sum5 = round(i['sum5'],3)
                sum6 = round(i['sum6'],3)
                sum7 = round(i['sum7'],3)
                sum8 = round(i['sum8'],3)
                sum9 = round(i['sum9'],3)
                sum10 = round(i['sum10'],3)
                sum11 = round(i['sum11'],3)
                sum12 = round(i['sum12'],3)
                fecfin =i['fmx']
                fecini =i['fmi']
                diffec = fecfin - fecini
                diffectxt = str(diffec)
                #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                # Crear las listas individuales para luego crear la general
                colu1t2.append(idp)
                colu2t2.append(prod)
                colu3t2.append(lot)
                colu4t2.append(sum1)
                colu5t2.append(sum2)
                colu6t2.append(sum3)
                colu7t2.append(sum4)
                colu8t2.append(sum5)
                colu9t2.append(sum6)
                colu10t2.append(sum7)
                colu11t2.append(sum8)
                colu12t2.append(sum9)
                colu13t2.append(sum10)
                colu14t2.append(sum11)
                colu15t2.append(sum12)
                colu16t2.append(fecfin)
                colu17t2.append(fecini)
                colu18t2.append(diffectxt)                   
        # Se crea un unica tabla
        resultado3 = zip(colu1t2,colu2t2,colu3t2,colu4t2,colu5t2,colu6t2,colu7t2,colu8t2,colu9t2,colu10t2,colu11t2,colu12t2,colu13t2,colu14t2,colu15t2,colu16t2,colu17t2,colu18t2)

        # Verifico si existen datos
        if list_mezct3:
            # Se crea una lista por producto por dia y se cuenta los SP y los PV               
            list_mezct3p = list_mezct3.values('id_producto','name_producto','lote').order_by('id_producto').annotate(ct=Count('id_producto'),
            fmx=Max('datetime_plc'),fmi=Min('datetime_plc'),
            sum1=Sum('bz_agre1_sp'),sum2=Sum('bz_agre1_pv'),sum3=Sum('bz_agre2_sp'),sum4=Sum('bz_agre2_pv'),
            sum5=Sum('bz_agre3_sp'),sum6=Sum('bz_agre3_pv'),sum7=Sum('bz_agre4_sp'),sum8=Sum('bz_agre4_pv'),
            sum9=Sum('bz_cem1_sp'),sum10=Sum('bz_cem1_pv'),sum11=Sum('bz_cem2_sp'),sum12=Sum('bz_cem2_pv'))
            list_mezct3p = list_mezct3p.order_by('lote')
            for i in list_mezct3p:
                #print(i)
                #print('id_prod: ',i['id_producto'])
                id = i['id_producto']
                idp = productos.objects.get(id = id)
                prod = i['name_producto']
                lot = i['lote']
                sum1 = round(i['sum1'],3)
                sum2 = round(i['sum2'],3)
                sum3 = round(i['sum3'],3)
                sum4 = round(i['sum4'],3)
                sum5 = round(i['sum5'],3)
                sum6 = round(i['sum6'],3)
                sum7 = round(i['sum7'],3)
                sum8 = round(i['sum8'],3)
                sum9 = round(i['sum9'],3)
                sum10 = round(i['sum10'],3)
                sum11 = round(i['sum11'],3)
                sum12 = round(i['sum12'],3)
                fecfin =i['fmx']
                fecini =i['fmi']
                diffec = fecfin - fecini
                diffectxt = str(diffec)
                #print(id,idp,prod,lot,sum1,sum2,sum3,sum4,sum5,sum6,sum7,sum8,sum9,sum10,sum11,sum12,fecfin,fecini,diffectxt)
                # Crear las listas individuales para luego crear la general
                colu1t3.append(idp)
                colu2t3.append(prod)
                colu3t3.append(lot)
                colu4t3.append(sum1)
                colu5t3.append(sum2)
                colu6t3.append(sum3)
                colu7t3.append(sum4)
                colu8t3.append(sum5)
                colu9t3.append(sum6)
                colu10t3.append(sum7)
                colu11t3.append(sum8)
                colu12t3.append(sum9)
                colu13t3.append(sum10)
                colu14t3.append(sum11)
                colu15t3.append(sum12)
                colu16t3.append(fecfin)
                colu17t3.append(fecini)
                colu18t3.append(diffectxt)                   
        # Se crea un unica tabla
        resultado4 = zip(colu1t3,colu2t3,colu3t3,colu4t3,colu5t3,colu6t3,colu7t3,colu8t3,colu9t3,colu10t3,colu11t3,colu12t3,colu13t3,colu14t3,colu15t3,colu16t3,colu17t3,colu18t3)

        # Ejecuto el Metodo de carga en Excel    
        excel(fecha0) 
        if excelE:
            if excelE == 1:
                env_excel = 0
            elif excelE > 1:
                env_excel = 1
        else:
            env_excel = 1

        greeting = {}

        greeting['heading'] = "PROCESO DE MEZCLADO POR DIA"
        greeting['pageview'] = "FECHA : "+ fech
        greeting['fechatxt'] = fech
        greeting['no_fila1'] = len(colu1) + 1
        greeting['no_fila2'] = len(colu1t1) + 1
        greeting['no_fila3'] = len(colu1t2) + 1
        greeting['no_fila4'] = len(colu1t3) + 1        
        greeting['sub_txt1'] = "PROCESO DE MEZCLADO"
        greeting['sub_txt2'] = "Reporte de KG # "
        greeting['table1'] = resultado1
        greeting['table2'] = resultado2
        greeting['table3'] = resultado3
        greeting['table4'] = resultado4
        greeting['table5'] = resultado_suma
        greeting['table6'] = resultado_silo         
        greeting['vfech'] = 2
        greeting['excelE'] = env_excel

        return render (request,'boquillas/inf_mezclado.html',greeting)        